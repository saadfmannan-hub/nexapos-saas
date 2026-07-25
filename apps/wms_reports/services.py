"""Presentation services and safe Excel export for WMS reports."""

from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils.text import slugify

from apps.core.date_ranges import business_localtime
from apps.wms_attendance.models import WmsAttendance

FORMULA_PREFIXES = ("=", "+", "-", "@")
ZERO = Decimal("0")


def duration_label(minutes):
    minutes = max(int(minutes or 0), 0)
    hours, remaining = divmod(minutes, 60)
    if hours and remaining:
        return f"{hours}h {remaining}m"
    if hours:
        return f"{hours}h"
    return f"{remaining}m"


def _time_label(value):
    return value.strftime("%H:%M") if value is not None else "—"


def _paired_times(first, second):
    return f"{_time_label(first)} / {_time_label(second)}"


def overall_attendance_status(record):
    """Summarize persisted Phase 3 shift statuses without recalculating them."""

    statuses = {record.morning_status, record.evening_status}
    if statuses == {WmsAttendance.Status.ABSENT}:
        return WmsAttendance.Status.ABSENT
    if WmsAttendance.Status.LATE in statuses:
        return WmsAttendance.Status.LATE
    return WmsAttendance.Status.PRESENT


def _is_incomplete(record):
    return any(
        (time_in is None) != (time_out is None)
        for time_in, time_out in (
            (record.morning_time_in, record.morning_time_out),
            (record.evening_time_in, record.evening_time_out),
        )
    )


def attendance_report(records):
    rows = []
    totals = {
        "present": 0,
        "late": 0,
        "absent": 0,
        "worked_minutes": 0,
        "missing_minutes": 0,
        "incomplete": 0,
    }
    for record in records:
        status = overall_attendance_status(record)
        incomplete = _is_incomplete(record)
        totals[status] += 1
        totals["worked_minutes"] += record.worked_minutes
        totals["missing_minutes"] += record.missing_minutes
        totals["incomplete"] += int(incomplete)
        rows.append(
            {
                "date": record.attendance_date,
                "employee_code": record.employee.employee_code,
                "employee_name": record.employee.full_name,
                "location_name": record.location.branch.name,
                "time_in": _paired_times(
                    record.morning_time_in,
                    record.evening_time_in,
                ),
                "time_out": _paired_times(
                    record.morning_time_out,
                    record.evening_time_out,
                ),
                "status": status,
                "status_label": dict(WmsAttendance.Status.choices)[status],
                "present": status == WmsAttendance.Status.PRESENT,
                "late": status == WmsAttendance.Status.LATE,
                "absent": status == WmsAttendance.Status.ABSENT,
                "worked_minutes": record.worked_minutes,
                "worked_label": duration_label(record.worked_minutes),
                "required_minutes": record.scheduled_minutes,
                "required_label": duration_label(record.scheduled_minutes),
                "missing_minutes": record.missing_minutes,
                "missing_label": duration_label(record.missing_minutes),
                "incomplete": incomplete,
                "indicator": (
                    "Incomplete check-in/check-out"
                    if incomplete
                    else (
                        f"Missing {duration_label(record.missing_minutes)}"
                        if record.missing_minutes
                        else "Complete"
                    )
                ),
            }
        )
    totals["worked_label"] = duration_label(totals["worked_minutes"])
    totals["missing_label"] = duration_label(totals["missing_minutes"])
    return {"rows": rows, "totals": totals}


def salary_report(records, business):
    rows = []
    totals = {
        "base_salary": ZERO,
        "eligible_pieces": 0,
        "piece_earnings": ZERO,
        "final_salary": ZERO,
    }
    for salary in records:
        is_fixed = salary.compensation_type_snapshot == "fixed_salary"
        base_salary = salary.fixed_monthly_salary_snapshot if is_fixed else ZERO
        piece_earnings = salary.gross_salary if not is_fixed else ZERO
        calculated_at = business_localtime(
            business,
            value=salary.calculated_at,
        )
        rows.append(
            {
                "employee_code": salary.employee_code_snapshot,
                "employee_name": salary.employee_name_snapshot,
                "salary_type": salary.get_compensation_type_snapshot_display(),
                "base_salary": base_salary,
                "eligible_pieces": salary.total_eligible_quantity,
                "piece_earnings": piece_earnings,
                "final_salary": salary.gross_salary,
                "status": salary.get_status_display(),
                "status_code": salary.status,
                "calculated_at": calculated_at,
                "public_id": salary.public_id,
            }
        )
        totals["base_salary"] += base_salary
        totals["eligible_pieces"] += salary.total_eligible_quantity
        totals["piece_earnings"] += piece_earnings
        totals["final_salary"] += salary.gross_salary
    return {"rows": rows, "totals": totals}


def _formula_safe(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def export_xlsx(*, report, business, metadata):
    """Create a self-describing, formula-safe Excel workbook."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report["sheet_name"][:31]
    column_count = max(len(report["columns"]), 1)

    sheet.cell(row=1, column=1, value=_formula_safe(report["title"]))
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=column_count,
    )
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0F172A")
    title_cell.alignment = Alignment(horizontal="center")

    metadata_rows = [
        ("Business", business.name),
        *metadata,
    ]
    row_number = 2
    for label, value in metadata_rows:
        sheet.cell(row=row_number, column=1, value=_formula_safe(label)).font = Font(bold=True)
        sheet.cell(
            row=row_number,
            column=2,
            value=_formula_safe(value),
        )
        row_number += 1
    row_number += 1
    header_row = row_number

    for column, value in enumerate(report["columns"], start=1):
        cell = sheet.cell(
            row=header_row,
            column=column,
            value=_formula_safe(value),
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center")

    for values in report["export_rows"]:
        row_number += 1
        for column, value in enumerate(values, start=1):
            sheet.cell(
                row=row_number,
                column=column,
                value=_formula_safe(value),
            )
    if report.get("export_totals"):
        row_number += 1
        for column, value in enumerate(report["export_totals"], start=1):
            cell = sheet.cell(
                row=row_number,
                column=column,
                value=_formula_safe(value),
            )
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for index, number_format in report.get("column_formats", {}).items():
        for row in range(header_row + 1, row_number + 1):
            sheet.cell(row=row, column=index + 1).number_format = number_format

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header_row}:{sheet.cell(row=header_row, column=column_count).column_letter}{row_number}"
    )
    for column in range(1, column_count + 1):
        values = [sheet.cell(row=row, column=column).value for row in range(1, row_number + 1)]
        width = min(
            max((len(str(value)) for value in values if value is not None), default=10) + 2,
            40,
        )
        sheet.column_dimensions[
            sheet.cell(row=header_row, column=column).column_letter
        ].width = max(width, 12)

    safe_name = slugify(report["filename"]) or "wms-report"
    response = HttpResponse(
        content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    )
    response["Content-Disposition"] = f'attachment; filename="{safe_name[:100]}.xlsx"'
    workbook.save(response)
    return response
