import csv
import io

from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render

from apps.audit import services as audit
from apps.core.decorators import require_permission
from apps.core.mixins import get_tenant_object
from apps.core.money import D
from apps.inventory import services as inventory
from apps.subscriptions import services as subscriptions

from .forms import (
    BrandForm,
    CategoryForm,
    ProductForm,
    ProductImportForm,
    TaxRateForm,
    UnitForm,
    VariantForm,
)
from .models import Brand, Category, Product, ProductVariant, TaxRate, Unit


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------
@require_permission("products.view")
def product_list(request):
    qs = (
        Product.objects.for_business(request.business)
        .select_related("category", "brand", "unit", "tax_rate")
        .filter(is_archived=False)
    )
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q) |
                       Q(barcode__icontains=q) | Q(internal_code__icontains=q))
    category_id = request.GET.get("category", "")
    if category_id.isdigit():
        qs = qs.filter(category_id=category_id)
    status = request.GET.get("status", "")
    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)
    sort = request.GET.get("sort", "name")
    if sort in ("name", "-name", "sale_price", "-sale_price", "-created_at"):
        qs = qs.order_by(sort)

    from django.core.paginator import Paginator

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Attach total stock to page items (single query)
    stock = {
        row["product_id"]: row["total"]
        for row in inventory.StockLevel.objects.for_business(request.business)
        .filter(product__in=[p.pk for p in page_obj])
        .values("product_id").annotate(total=Sum("quantity"))
    }
    for p in page_obj:
        p.total_stock = stock.get(p.pk, 0)

    categories = Category.objects.for_business(request.business).filter(is_active=True)
    p_cur, p_lim, _ = subscriptions.limit_state(request.business, "products")
    return render(request, "catalog/product_list.html", {
        "page_obj": page_obj, "q": q, "categories": categories,
        "active_nav": "products", "product_count": p_cur, "product_limit": p_lim,
        "querystring": _qs_without_page(request),
    })


def _qs_without_page(request):
    params = request.GET.copy()
    params.pop("page", None)
    encoded = params.urlencode()
    return f"{encoded}&" if encoded else ""


@require_permission("products.manage")
def product_form(request, public_id=None):
    instance = None
    if public_id:
        instance = get_tenant_object(Product, request.business, public_id=public_id)
    else:
        from apps.subscriptions.helpers import guard_limit

        blocked = guard_limit(request, "products")
        if blocked:
            return blocked

    form = ProductForm(request.business, request.POST or None,
                       request.FILES or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.business = request.business
        product.save()
        opening = form.cleaned_data.get("opening_stock")
        warehouse = form.cleaned_data.get("opening_warehouse")
        if not public_id and opening and warehouse and product.is_stocked:
            inventory.set_opening_stock(
                business=request.business, warehouse=warehouse, product=product,
                quantity=opening, unit_cost=product.purchase_price,
                user=request.user,
            )
        audit.log("product.saved", request=request, module="catalog", obj=product,
                  description=f"Product '{product.name}' saved.")
        messages.success(request, "Product saved.")
        if product.has_variants:
            return redirect("catalog:product_detail", public_id=product.public_id)
        return redirect("catalog:product_list")
    return render(request, "catalog/product_form.html",
                  {"form": form, "product": instance, "active_nav": "products"})


@require_permission("products.view")
def product_detail(request, public_id):
    product = get_tenant_object(
        Product.objects.select_related("category", "brand", "unit", "tax_rate"),
        request.business, public_id=public_id,
    )
    variants = product.variants.all()
    levels = (
        inventory.StockLevel.objects.for_business(request.business)
        .filter(product=product).select_related("warehouse", "variant")
    )
    movements = (
        inventory.StockMovement.objects.for_business(request.business)
        .filter(product=product).select_related("warehouse", "user")[:30]
    )
    show_cost = request.membership.has_perm("cost.view")
    return render(request, "catalog/product_detail.html", {
        "product": product, "variants": variants, "levels": levels,
        "movements": movements, "active_nav": "products", "show_cost": show_cost,
    })


@require_permission("products.manage")
def product_archive(request, public_id):
    product = get_tenant_object(Product, request.business, public_id=public_id)
    if request.method == "POST":
        # Products referenced by invoices are archived, never deleted.
        product.is_archived = True
        product.is_active = False
        product.save(update_fields=["is_archived", "is_active"])
        audit.log("product.archived", request=request, module="catalog", obj=product,
                  description=f"Product '{product.name}' archived.")
        messages.success(request, f"'{product.name}' archived.")
    return redirect("catalog:product_list")


@require_permission("products.manage")
def variant_form(request, product_id, public_id=None):
    product = get_tenant_object(Product, request.business, public_id=product_id)
    instance = None
    if public_id:
        instance = get_tenant_object(ProductVariant, request.business, public_id=public_id)
        if instance.product_id != product.id:
            from django.http import Http404
            raise Http404
    form = VariantForm(request.business, request.POST or None,
                       request.FILES or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        variant = form.save(commit=False)
        variant.business = request.business
        variant.product = product
        variant.attributes = form.build_attributes()
        if not variant.name:
            variant.name = " / ".join(variant.attributes.values()) or "Variant"
        variant.save()
        if product.product_type != Product.Type.VARIANT:
            product.product_type = Product.Type.VARIANT
            product.save(update_fields=["product_type"])
        messages.success(request, "Variant saved.")
        return redirect("catalog:product_detail", public_id=product.public_id)
    return render(request, "catalog/variant_form.html",
                  {"form": form, "product": product, "variant": instance,
                   "active_nav": "products"})


# ---------------------------------------------------------------------------
# Barcode generation / labels
# ---------------------------------------------------------------------------
@require_permission("products.view")
def product_barcode_svg(request, public_id):
    """Server-generated Code128 barcode as SVG."""
    import barcode
    from barcode.writer import SVGWriter

    product = get_tenant_object(Product, request.business, public_id=public_id)
    code = product.barcode or product.sku or f"P{product.pk:08d}"
    buffer = io.BytesIO()
    barcode.get("code128", code, writer=SVGWriter()).write(buffer)
    return HttpResponse(buffer.getvalue(), content_type="image/svg+xml")


@require_permission("products.view")
def product_labels(request, public_id):
    product = get_tenant_object(Product, request.business, public_id=public_id)
    try:
        count = max(1, min(int(request.GET.get("count", 12)), 120))
    except ValueError:
        count = 12
    return render(request, "catalog/labels.html", {
        "product": product, "count_range": range(count),
        "business": request.business,
    })


# ---------------------------------------------------------------------------
# Categories / brands / units / taxes (combined setup screens)
# ---------------------------------------------------------------------------
def _simple_crud(request, model, form_class, list_template, name, perm="products.manage",
                 extra=None):
    @require_permission(perm)
    def handler(request):
        instance = None
        edit_id = request.GET.get("edit")
        if edit_id:
            instance = get_tenant_object(model, request.business, public_id=edit_id)
        form = form_class(request.business, request.POST or None, instance=instance)
        if request.method == "POST" and form.is_valid():
            obj = form.save(commit=False)
            obj.business = request.business
            obj.save()
            if isinstance(obj, TaxRate) and obj.is_default:
                TaxRate.objects.for_business(request.business).exclude(pk=obj.pk).update(
                    is_default=False
                )
            messages.success(request, f"{name} saved.")
            return redirect(request.path)
        items = model.objects.for_business(request.business)
        ctx = {"form": form, "items": items, "editing": instance,
               "active_nav": "catalog_setup"}
        if extra:
            ctx.update(extra(request))
        return render(request, list_template, ctx)

    return handler(request)


def category_list(request):
    return _simple_crud(request, Category, CategoryForm,
                        "catalog/category_list.html", "Category")


def brand_list(request):
    return _simple_crud(request, Brand, BrandForm, "catalog/brand_list.html", "Brand")


def unit_list(request):
    return _simple_crud(request, Unit, UnitForm, "catalog/unit_list.html", "Unit")


def tax_list(request):
    return _simple_crud(request, TaxRate, TaxRateForm, "catalog/tax_list.html",
                        "Tax rate", perm="settings.manage")


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
IMPORT_COLUMNS = ["name", "sku", "barcode", "category", "brand", "unit",
                  "purchase_price", "sale_price", "opening_stock", "reorder_level"]


@require_permission("products.import")
def product_import(request):
    form = ProductImportForm(request.POST or None, request.FILES or None)
    results = None
    if request.method == "POST" and form.is_valid():
        try:
            subscriptions.require_operational(request.business)
        except subscriptions.SubscriptionInactive as exc:
            messages.error(request, str(exc))
            return redirect("catalog:product_list")
        rows, parse_errors = _parse_import_file(form.cleaned_data["file"])
        if parse_errors:
            messages.error(request, parse_errors)
        else:
            results = _import_rows(request, rows, form.cleaned_data["match_by"])
            audit.log("products.imported", request=request, module="catalog",
                      description=(f"Product import: {results['created']} created, "
                                   f"{len(results['errors'])} errors, "
                                   f"{results['skipped']} skipped."))
    return render(request, "catalog/product_import.html",
                  {"form": form, "results": results, "columns": IMPORT_COLUMNS,
                   "active_nav": "products"})


@require_permission("products.import")
def import_template(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="product_import_template.csv"'
    writer = csv.writer(response)
    writer.writerow(IMPORT_COLUMNS)
    writer.writerow(["Example T-Shirt", "TSH-001", "6291041500213", "Clothing",
                     "Generic", "Piece", "2.500", "4.900", "10", "5"])
    return response


def _parse_import_file(file):
    name = file.name.lower()
    rows = []
    try:
        if name.endswith(".csv"):
            text = io.TextIOWrapper(file.file, encoding="utf-8-sig")
            reader = csv.DictReader(text)
            rows = [dict(r) for r in reader]
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook

            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2):
                values = [c.value for c in row]
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append(dict(zip(headers, values)))
        else:
            return [], "Unsupported file type. Upload .csv or .xlsx."
    except Exception as exc:  # parsing must never 500
        return [], f"Could not read file: {exc}"
    if len(rows) > 5000:
        return [], "File exceeds the 5000-row limit per import."
    return rows, None


def _import_rows(request, rows, match_by):
    from django.db import transaction

    business = request.business
    created, skipped, errors = 0, 0, []
    from apps.branches.models import Warehouse

    default_warehouse = (
        Warehouse.objects.for_business(business).filter(is_active=True)
        .order_by("-is_default").first()
    )
    for idx, row in enumerate(rows, start=2):
        norm = {str(k).strip().lower(): ("" if v is None else str(v).strip())
                for k, v in row.items() if k}
        name = norm.get("name", "")
        if not name:
            errors.append(f"Row {idx}: missing product name.")
            continue
        match_value = norm.get(match_by, "")
        if match_value and Product.objects.for_business(business).filter(
            **{match_by: match_value}
        ).exists():
            skipped += 1
            continue
        if not match_value and match_by == "name" and Product.objects.for_business(
            business
        ).filter(name__iexact=name).exists():
            skipped += 1
            continue
        try:
            _current, limit, allowed = subscriptions.limit_state(business, "products")
            if not allowed:
                errors.append(f"Row {idx}: plan product limit ({limit}) reached.")
                break
            with transaction.atomic():
                sku, bc = norm.get("sku", ""), norm.get("barcode", "")
                if sku and Product.objects.for_business(business).filter(sku=sku).exists():
                    raise ValueError(f"duplicate SKU {sku}")
                if bc and Product.objects.for_business(business).filter(barcode=bc).exists():
                    raise ValueError(f"duplicate barcode {bc}")
                category = None
                if norm.get("category"):
                    category, _ = Category.objects.get_or_create(
                        business=business, name=norm["category"], parent=None)
                brand = None
                if norm.get("brand"):
                    brand, _ = Brand.objects.get_or_create(business=business,
                                                           name=norm["brand"])
                unit = None
                if norm.get("unit"):
                    unit = Unit.objects.for_business(business).filter(
                        name__iexact=norm["unit"]).first()
                product = Product.objects.create(
                    business=business, name=name[:200], sku=sku[:60],
                    barcode=bc[:80], category=category, brand=brand, unit=unit,
                    purchase_price=D(norm.get("purchase_price")),
                    sale_price=D(norm.get("sale_price")),
                    reorder_level=D(norm.get("reorder_level")),
                )
                opening = D(norm.get("opening_stock"))
                if opening > 0 and default_warehouse:
                    inventory.set_opening_stock(
                        business=business, warehouse=default_warehouse,
                        product=product, quantity=opening,
                        unit_cost=product.purchase_price, user=request.user,
                    )
            created += 1
        except Exception as exc:
            errors.append(f"Row {idx}: {exc}")
    return {"created": created, "skipped": skipped, "errors": errors,
            "total": len(rows)}
