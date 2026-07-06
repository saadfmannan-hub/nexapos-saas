"""Reusable Excel/CSV import parsing and error-report helpers.

Shared by the customer, product and inventory importers so file parsing,
row limits and the downloadable error report are implemented once.
"""
import csv
import io

from django.http import HttpResponse

MAX_IMPORT_ROWS = 10000


def parse_tabular_file(uploaded_file, max_rows=MAX_IMPORT_ROWS):
    """Parse an uploaded .csv or .xlsx into a list of dict rows.

    Returns (rows, error). Header keys are lower-cased and stripped.
    Never raises — parsing failures come back as a friendly error string.
    """
    name = (uploaded_file.name or "").lower()
    rows = []
    try:
        if name.endswith(".csv"):
            text = io.TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
            reader = csv.DictReader(text)
            for raw in reader:
                rows.append({(k or "").strip().lower(): v for k, v in raw.items()})
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook

            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(max_row=1), None)
            if header_row is None:
                return [], "The file is empty."
            headers = [str(c.value or "").strip().lower() for c in header_row]
            for row in ws.iter_rows(min_row=2):
                values = [c.value for c in row]
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append(dict(zip(headers, values, strict=False)))
        else:
            return [], "Unsupported file type. Upload a .csv or .xlsx file."
    except Exception as exc:  # parsing must never 500
        return [], f"Could not read the file: {exc}"
    if len(rows) > max_rows:
        return [], f"File exceeds the {max_rows:,}-row limit per import."
    return rows, None


def normalize_row(row):
    """Lower-case keys, string-coerce values, strip whitespace."""
    return {
        str(k).strip().lower(): ("" if v is None else str(v).strip())
        for k, v in row.items() if k is not None
    }


def error_report_response(filename, errors):
    """Build a downloadable CSV error report: Row, Field, Error."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Row", "Field", "Error"])
    for error in errors:
        if len(error) == 3:
            row_no, field, message = error
        else:
            row_no, message = error
            field = ""
        writer.writerow([row_no, field, message])
    return response
