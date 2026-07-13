"""CSV / Excel / PDF export builders for report datasets."""
import csv
from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse

from .pdf import render_pdf


def _cell(value):
    if value is None or value == "":
        return "-"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return str(value)
    return value


def _csv_cell(value):
    return "-" if value is None or value == "" else value


def export_csv(title, data):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{title}.csv"'
    writer = csv.writer(response)
    writer.writerow(data["columns"])
    for row in data["rows"]:
        writer.writerow([_csv_cell(v) for v in row])
    if data.get("totals"):
        writer.writerow(data["totals"])
    if data.get("summary"):
        writer.writerow([])
        for label, value in data["summary"]:
            writer.writerow([label, _csv_cell(value)])
    return response


def export_xlsx(title, data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    ws.append(data["columns"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in data["rows"]:
        ws.append([_cell(v) for v in row])
    if data.get("totals"):
        ws.append([_cell(v) for v in data["totals"]])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    if data.get("summary"):
        ws.append([])
        for label, value in data["summary"]:
            ws.append([label, _cell(value)])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for idx, col in enumerate(data["columns"], start=1):
        width = max(len(str(col)) + 2, 12)
        if idx == 1 and data.get("summary"):
            width = max(width, max(len(str(label)) + 2 for label, _ in data["summary"]))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    wb.save(response)
    return response


def export_pdf(title, data, business, filters_label=""):
    pdf = render_pdf("reports/report_pdf.html", {
        "title": title, "data": data, "business": business,
        "filters_label": filters_label,
    })
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title}.pdf"'
    return response
