import csv
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import RequestFactory
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.branches.models import Branch, Warehouse
from apps.catalog.models import Product
from apps.inventory import services as inventory
from apps.reports.queries import sales_detailed
from apps.sales import services as sales
from apps.sales.models import PaymentMethod
from apps.sales.views import _invoice_context, _job_card_context

from .base import TenantTestCase

D = Decimal


class GarmentClassificationVisibilityTests(TenantTestCase):
    def setUp(self):
        self.allow_no_shift()
        self.product_a.is_tailoring_item = True
        self.product_a.save(update_fields=["is_tailoring_item"])
        self.client.force_login(self.owner_a)

    def tailoring_sale(self, lines):
        return self.make_sale(
            items=[{
                "product": self.product_a,
                "quantity": D(str(quantity)),
                "unit_price": self.product_a.sale_price,
                "garment_classification": classification,
                "collection_type": "normal",
                "tailoring_details": {
                    "customer_notes": f"{classification} fitting",
                },
            } for classification, quantity in lines],
            delivery_date=timezone.localdate(),
        )

    def complete_sale_at(self, *, product, classification, branch, warehouse):
        line = {
            "product": product,
            "quantity": D("1"),
            "unit_price": product.sale_price,
            "garment_classification": classification,
            "collection_type": "normal",
        }
        total = sales.compute_line(
            product,
            None,
            line["quantity"],
            line["unit_price"],
            D("0"),
            self.business_a.settings.prices_include_tax,
        )["total"]
        return sales.complete_sale(
            business=self.business_a,
            branch=branch,
            warehouse=warehouse,
            cashier=self.owner_a,
            customer=self.walk_in_a,
            membership=self.membership_a(),
            items=[line],
            payments=[{"method": self.cash_a, "amount": total}],
            delivery_date=timezone.localdate(),
        )

    def render_job_card(self, sale):
        item = sale.items.select_related("product__unit", "variant").get()
        request = RequestFactory().get("/")
        request.business = self.business_a
        return render_to_string(
            "invoices/workshop_job_card.html",
            _job_card_context(sale, request, [item], sale_item=item),
        )

    def historical_tailoring_sale(self):
        self.product_a.is_tailoring_item = False
        self.product_a.save(update_fields=["is_tailoring_item"])
        sale = self.make_sale(items=[{
            "product": self.product_a,
            "quantity": D("2"),
            "unit_price": self.product_a.sale_price,
        }])
        self.product_a.is_tailoring_item = True
        self.product_a.save(update_fields=["is_tailoring_item"])
        return sale

    def test_adult_is_visible_on_job_card(self):
        html = self.render_job_card(self.tailoring_sale([("adult", 1)]))
        self.assertIn('<div class="label">Garment</div>', html)
        self.assertIn('<div class="value">Adult</div>', html)

    def test_child_is_visible_on_job_card(self):
        html = self.render_job_card(self.tailoring_sale([("child", 1)]))
        self.assertIn('<div class="value">Child</div>', html)

    def test_mixed_invoice_job_cards_keep_line_classifications(self):
        sale = self.tailoring_sale([("adult", 1), ("child", 1)])
        with patch("apps.reports.pdf.render_pdf", return_value=b"%PDF fake") as render_pdf:
            response = self.client.get(
                reverse("sales:workshop_job_card_pdf", args=[sale.public_id])
            )
        self.assertEqual(response.status_code, 200)
        cards = render_pdf.call_args.args[1]["job_cards"]
        self.assertEqual(
            [card["job_item"].garment_classification for card in cards],
            ["adult", "child"],
        )
        html = render_to_string("invoices/workshop_job_card.html", {
            "job_cards": cards,
        })
        self.assertEqual(html.count('<div class="label">Garment</div>'), 2)
        self.assertIn('<div class="value">Adult</div>', html)
        self.assertIn('<div class="value">Child</div>', html)

    def test_adult_is_visible_on_sale_detail(self):
        sale = self.tailoring_sale([("adult", 1)])
        html = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        ).content.decode()
        self.assertIn(
            'Garment: <span class="fw-semibold text-body">Adult</span>', html
        )

    def test_child_is_visible_on_sale_detail(self):
        sale = self.tailoring_sale([("child", 1)])
        html = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        ).content.decode()
        self.assertIn(
            'Garment: <span class="fw-semibold text-body">Child</span>', html
        )

    def test_historical_blank_classification_is_readable(self):
        sale = self.historical_tailoring_sale()
        item = sale.items.select_related("product").get()
        self.assertEqual(item.garment_classification, "")
        self.assertEqual(item.garment_classification_label, "Legacy / Not Recorded")
        response = self.client.get(reverse("sales:detail", args=[sale.public_id]))
        self.assertContains(response, "Legacy / Not Recorded")

    def test_sales_report_has_line_classification_and_piece_totals(self):
        self.tailoring_sale([("adult", 2), ("child", 3)])
        response = self.client.get(reverse("reports:view", args=["sales_detailed"]))
        data = response.context["data"]
        self.assertIn("Garment Classification", data["columns"])
        self.assertEqual([row[6] for row in data["rows"]], ["Adult", "Child"])
        self.assertEqual([row[8] for row in data["rows"]], [D("2"), D("3")])
        summary = dict(data["summary"])
        self.assertEqual(summary["Total Adult Pieces"], D("2"))
        self.assertEqual(summary["Total Child Pieces"], D("3"))
        self.assertEqual(summary["Total Legacy/Unclassified Pieces"], D("0"))

    def test_sales_report_counts_legacy_tailoring_without_counting_retail(self):
        self.historical_tailoring_sale()
        retail = Product.objects.create(
            business=self.business_a,
            name="Retail Accessory",
            sku="RET-ONLY",
            purchase_price=D("1"),
            sale_price=D("2"),
        )
        inventory.set_opening_stock(
            business=self.business_a,
            warehouse=self.warehouse_a,
            product=retail,
            quantity=D("10"),
            unit_cost=D("1"),
            user=self.owner_a,
        )
        self.make_sale(items=[{
            "product": retail,
            "quantity": D("4"),
            "unit_price": retail.sale_price,
        }])
        data = sales_detailed(self.business_a, {})
        labels = {row[5]: row[6] for row in data["rows"]}
        self.assertEqual(labels[self.product_a.name], "Legacy / Not Recorded")
        self.assertEqual(labels[retail.name], "Not Applicable")
        self.assertEqual(
            dict(data["summary"])["Total Legacy/Unclassified Pieces"], D("2")
        )

    def test_sales_report_filters_classification_product_date_and_branch(self):
        self.tailoring_sale([("adult", 1), ("child", 1)])
        other_product = Product.objects.create(
            business=self.business_a,
            name="Other Tailoring Product",
            sku="TAIL-OTHER",
            purchase_price=D("4"),
            sale_price=D("12"),
            tax_rate=self.tax_a,
            is_tailoring_item=True,
            estimated_adult_fabric=D("3.500"),
            estimated_child_fabric=D("2.250"),
        )
        inventory.set_opening_stock(
            business=self.business_a,
            warehouse=self.warehouse_a,
            product=other_product,
            quantity=D("10"),
            unit_cost=D("4"),
            user=self.owner_a,
        )
        self.complete_sale_at(
            product=other_product,
            classification="adult",
            branch=self.branch_a,
            warehouse=self.warehouse_a,
        )
        second_branch = Branch.objects.create(
            business=self.business_a,
            name="Second Branch",
            code="SECOND-REPORT",
        )
        second_warehouse = Warehouse.objects.create(
            business=self.business_a,
            branch=second_branch,
            name="Second Report Warehouse",
            code="SECOND-REPORT-WH",
        )
        inventory.set_opening_stock(
            business=self.business_a,
            warehouse=second_warehouse,
            product=self.product_a,
            quantity=D("10"),
            unit_cost=D("4"),
            user=self.owner_a,
        )
        branch_sale = self.complete_sale_at(
            product=self.product_a,
            classification="child",
            branch=second_branch,
            warehouse=second_warehouse,
        )
        url = reverse("reports:view", args=["sales_detailed"])

        adult = self.client.get(url, {"garment_classification": "adult"})
        self.assertTrue(adult.context["data"]["rows"])
        self.assertTrue(all(row[6] == "Adult" for row in adult.context["data"]["rows"]))

        product = self.client.get(url, {"product": other_product.id})
        self.assertEqual({row[5] for row in product.context["data"]["rows"]}, {
            other_product.name,
        })

        branch = self.client.get(url, {"branch": second_branch.id})
        self.assertEqual([row[0] for row in branch.context["data"]["rows"]], [
            branch_sale.invoice_number,
        ])

        old_date = self.client.get(url, {
            "from": "1990-01-01",
            "to": "1990-01-02",
        })
        self.assertEqual(old_date.context["data"]["rows"], [])

    def test_csv_export_contains_mixed_classification_and_summary(self):
        self.tailoring_sale([("adult", 1), ("child", 2)])
        response = self.client.get(
            reverse("reports:view", args=["sales_detailed"]),
            {"export": "csv"},
        )
        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertIn("Garment Classification", rows[0])
        classification_index = rows[0].index("Garment Classification")
        self.assertEqual(
            {row[classification_index] for row in rows[1:3]},
            {"Adult", "Child"},
        )
        self.assertIn(["Total Adult Pieces", "1.000"], rows)
        self.assertIn(["Total Child Pieces", "2.000"], rows)

    def test_xlsx_export_contains_mixed_classification_and_summary(self):
        self.tailoring_sale([("adult", 1), ("child", 2)])
        response = self.client.get(
            reverse("reports:view", args=["sales_detailed"]),
            {"export": "xlsx"},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        header = list(rows[0])
        classification_index = header.index("Garment Classification")
        self.assertEqual(
            {rows[1][classification_index], rows[2][classification_index]},
            {"Adult", "Child"},
        )
        summary = {row[0]: row[1] for row in rows if row[0] and str(row[0]).startswith("Total ")}
        self.assertEqual(summary["Total Adult Pieces"], 1)
        self.assertEqual(summary["Total Child Pieces"], 2)

    def test_a4_invoice_renders_each_line_classification(self):
        sale = self.tailoring_sale([("adult", 1), ("child", 1)])
        response = self.client.get(reverse("sales:invoice", args=[sale.public_id]))
        self.assertContains(response, "Garment: Adult")
        self.assertContains(response, "Garment: Child")

    def test_thermal_receipts_render_compact_line_classification(self):
        sale = self.tailoring_sale([("adult", 1), ("child", 1)])
        receipt_80 = self.client.get(reverse("sales:receipt", args=[sale.public_id]))
        self.assertContains(receipt_80, "Garment: Adult")
        self.assertContains(receipt_80, "Garment: Child")

        receipt_58 = render_to_string(
            "invoices/receipt_58mm.html",
            _invoice_context(sale),
        )
        self.assertIn("G: Adult", receipt_58)
        self.assertIn("G: Child", receipt_58)

    def test_report_and_export_remain_tenant_and_branch_scoped(self):
        self.product_b.is_tailoring_item = True
        self.product_b.save(update_fields=["is_tailoring_item"])
        self.allow_no_shift(self.business_b)
        cash_b = PaymentMethod.objects.for_business(self.business_b).get(kind="cash")
        sales.complete_sale(
            business=self.business_b,
            branch=self.branch_b,
            warehouse=self.warehouse_b,
            cashier=self.owner_b,
            customer=self.walk_in_b,
            membership=self.business_b.memberships.get(user=self.owner_b),
            items=[{
                "product": self.product_b,
                "quantity": D("1"),
                "unit_price": self.product_b.sale_price,
                "garment_classification": "child",
            }],
            payments=[{"method": cash_b, "amount": D("5")}],
            delivery_date=timezone.localdate(),
        )
        url = reverse("reports:view", args=["sales_detailed"])
        response = self.client.get(url)
        self.assertNotIn(self.product_b.name, [row[5] for row in response.context["data"]["rows"]])
        wrong_branch = self.client.get(url, {"branch": self.branch_b.id})
        self.assertEqual(wrong_branch.context["data"]["rows"], [])
        csv_response = self.client.get(url, {"export": "csv"})
        self.assertNotIn(self.product_b.name, csv_response.content.decode("utf-8"))
