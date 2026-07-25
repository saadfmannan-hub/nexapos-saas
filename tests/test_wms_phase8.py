"""Focused regression coverage for WMS Phase 8 reports."""

from datetime import UTC, date, datetime, time
from io import BytesIO
from unittest.mock import patch

from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from openpyxl import load_workbook

from apps.wms_alterations.models import WmsAlteration
from apps.wms_attendance import services as attendance_services
from apps.wms_production import services as production_services
from apps.wms_reports import selectors
from apps.wms_salary.models import WmsSalary
from apps.wms_workforce.models import WmsEmployeeCategoryAssignment
from tests.test_wms_phase2 import make_category, make_employee
from tests.test_wms_phase7 import WmsPhase7Base


class WmsPhase8ReportTests(WmsPhase7Base):
    report_date = date(2026, 7, 12)

    def report_url(self, name, **params):
        return reverse(f"wms:{name}"), params

    def create_report_production(
        self,
        *,
        employee=None,
        production_date=None,
        daily_total=9,
        override_quantity=3,
        default_quantity=4,
    ):
        return self.create_production(
            employee=employee,
            production_date=production_date or self.report_date,
            daily_total=daily_total,
            override_quantity=override_quantity,
            default_quantity=default_quantity,
        )

    def create_report_attendance(
        self,
        attendance_date,
        *,
        employee=None,
        morning_in=time(10, 0),
        morning_out=time(13, 0),
        evening_in=time(16, 30),
        evening_out=time(22, 0),
    ):
        employee = employee or self.fixed_employee
        return attendance_services.create_attendance(
            business=employee.business,
            employee=employee,
            attendance_date=attendance_date,
            time_values={
                "morning_time_in": morning_in,
                "morning_time_out": morning_out,
                "evening_time_in": evening_in,
                "evening_time_out": evening_out,
            },
            user=employee.business.owner,
        )

    def test_report_pages_and_navigation_are_available_to_owner(self):
        page_names = (
            "report_index",
            "report_daily_production",
            "report_monthly_production",
            "report_attendance_summary",
            "report_individual_attendance",
            "report_salary",
        )
        for page_name in page_names:
            with self.subTest(page_name=page_name):
                response = self.client.get(reverse(f"wms:{page_name}"))
                self.assertEqual(response.status_code, 200)
        response = self.client.get(reverse("wms:report_index"))
        self.assertContains(response, "Daily Finished Pieces")
        self.assertContains(response, "Monthly Salary")
        self.assertContains(response, reverse("wms:report_salary"))

    def test_daily_report_uses_stored_total_and_category_breakdown(self):
        self.create_report_production()
        self.create_report_production(
            production_date=date(2026, 7, 13),
            daily_total=20,
            override_quantity=8,
            default_quantity=7,
        )
        response = self.client.get(
            reverse("wms:report_daily_production"),
            {"report_date": "2026-07-12"},
        )
        self.assertEqual(response.status_code, 200)
        report = response.context["report"]
        self.assertEqual(report["grand_total"], 9)
        self.assertEqual(len(report["rows"]), 1)
        self.assertEqual(report["rows"][0]["total"], 9)
        totals = {
            category["name"]: report["category_totals"][index]
            for index, category in enumerate(report["categories"])
        }
        self.assertEqual(
            totals,
            {"Default Category": 4, "Override Category": 3},
        )
        self.assertNotContains(response, "20</td>", html=False)

        filtered = self.client.get(
            reverse("wms:report_daily_production"),
            {
                "report_date": "2026-07-12",
                "category": str(self.category_override.public_id),
                "employee": str(self.piece_employee.public_id),
                "location": str(self.location_a1.public_id),
            },
        )
        filtered_report = filtered.context["report"]
        self.assertEqual(
            [item["name"] for item in filtered_report["categories"]],
            ["Override Category"],
        )
        self.assertEqual(filtered_report["category_totals"], [3])
        self.assertEqual(filtered_report["grand_total"], 9)

    def test_monthly_production_has_every_day_and_assignment_categories(self):
        self.create_report_production()
        response = self.client.get(
            reverse("wms:report_monthly_production"),
            {
                "report_month": "2026-07",
                "employee": str(self.piece_employee.public_id),
            },
        )
        self.assertEqual(response.status_code, 200)
        report = response.context["report"]
        self.assertEqual(len(report["rows"]), 31)
        production_day = next(row for row in report["rows"] if row["date"] == self.report_date)
        self.assertEqual(production_day["total"], 9)
        self.assertEqual(report["grand_total"], 9)
        totals = {
            category["name"]: report["category_totals"][index]
            for index, category in enumerate(report["categories"])
        }
        self.assertEqual(
            totals,
            {"Default Category": 4, "Override Category": 3},
        )

        unused = make_category(self.business_a, "Unassigned", "UNASSIGNED")
        response = self.client.get(
            reverse("wms:report_monthly_production"),
            {
                "report_month": "2026-07",
                "employee": str(self.piece_employee.public_id),
            },
        )
        self.assertNotContains(response, unused.name)

    def test_alterations_never_contribute_to_production_reports(self):
        self.create_report_production(daily_total=11)
        WmsAlteration.objects.create(
            business=self.business_a,
            location=self.location_a1,
            original_order_reference="ORDER-ALTERATION",
            alteration_reference="ALT-001",
            reason=WmsAlteration.Reason.IRON,
            mistake_by=WmsAlteration.MistakeBy.CUSTOMER,
            assigned_employee=self.piece_employee,
            alteration_date=self.report_date,
            status=WmsAlteration.Status.OPEN,
            created_by=self.owner_a,
            updated_by=self.owner_a,
        )
        response = self.client.get(
            reverse("wms:report_daily_production"),
            {"report_date": self.report_date.isoformat()},
        )
        self.assertEqual(response.context["report"]["grand_total"], 11)
        self.assertNotContains(response, "ALT-001")

    def test_daily_default_uses_business_local_date(self):
        self.create_report_production()
        utc_time = datetime(2026, 7, 11, 20, 30, tzinfo=UTC)
        with patch(
            "apps.core.date_ranges.timezone.now",
            return_value=utc_time,
        ):
            response = self.client.get(reverse("wms:report_daily_production"))
        self.assertEqual(response.context["report"]["grand_total"], 9)
        self.assertEqual(
            response.context["report"]["period_label"],
            "July 12, 2026",
        )

    def test_attendance_summary_uses_persisted_phase3_results(self):
        self.create_report_attendance(date(2026, 7, 1))
        late = self.create_report_attendance(
            date(2026, 7, 2),
            morning_in=time(10, 30),
            morning_out=None,
            evening_in=time(16, 30),
            evening_out=time(22, 0),
        )
        absent = self.create_report_attendance(
            date(2026, 7, 3),
            morning_in=None,
            morning_out=None,
            evening_in=None,
            evening_out=None,
        )
        response = self.client.get(
            reverse("wms:report_attendance_summary"),
            {"date_from": "2026-07-01", "date_to": "2026-07-31"},
        )
        self.assertEqual(response.status_code, 200)
        report = response.context["report"]
        self.assertEqual(report["totals"]["present"], 1)
        self.assertEqual(report["totals"]["late"], 1)
        self.assertEqual(report["totals"]["absent"], 1)
        self.assertEqual(report["totals"]["incomplete"], 1)
        self.assertEqual(
            report["totals"]["worked_minutes"],
            510 + late.worked_minutes + absent.worked_minutes,
        )
        self.assertEqual(
            report["totals"]["missing_minutes"],
            late.missing_minutes + absent.missing_minutes,
        )
        self.assertContains(response, "Incomplete check-in/check-out")

    def test_individual_attendance_totals_are_employee_only(self):
        self.create_report_attendance(date(2026, 7, 1))
        self.create_report_attendance(
            date(2026, 7, 2),
            morning_in=None,
            morning_out=None,
            evening_in=None,
            evening_out=None,
        )
        other_employee = make_employee(
            self.business_a,
            self.location_a1,
            "P8-OTHER",
        )
        self.create_report_attendance(
            date(2026, 7, 1),
            employee=other_employee,
        )
        response = self.client.get(
            reverse("wms:report_individual_attendance"),
            {
                "report_month": "2026-07",
                "employee": str(self.fixed_employee.public_id),
            },
        )
        report = response.context["report"]
        self.assertEqual(len(report["rows"]), 2)
        self.assertEqual(report["totals"]["present"], 1)
        self.assertEqual(report["totals"]["absent"], 1)
        self.assertNotIn(
            other_employee.full_name,
            [row["employee_name"] for row in report["rows"]],
        )

    def test_salary_report_reuses_fixed_and_piece_calculation_snapshots(self):
        self.create_report_attendance(date(2026, 7, 10))
        self.create_report_production()
        fixed_salary = self.calculate(self.fixed_employee)
        piece_salary = self.calculate(self.piece_employee)
        original_state = list(
            WmsSalary.objects.order_by("pk").values_list(
                "pk",
                "status",
                "gross_salary",
                "calculated_at",
                "finalized_at",
            )
        )
        response = self.client.get(
            reverse("wms:report_salary"),
            {"report_month": "2026-07"},
        )
        self.assertEqual(response.status_code, 200)
        rows = {row["employee_code"]: row for row in response.context["report"]["rows"]}
        self.assertEqual(
            rows[self.fixed_employee.employee_code]["base_salary"],
            fixed_salary.fixed_monthly_salary_snapshot,
        )
        self.assertEqual(
            rows[self.piece_employee.employee_code]["eligible_pieces"],
            piece_salary.total_eligible_quantity,
        )
        self.assertEqual(
            rows[self.piece_employee.employee_code]["piece_earnings"],
            piece_salary.gross_salary,
        )
        self.assertEqual(
            original_state,
            list(
                WmsSalary.objects.order_by("pk").values_list(
                    "pk",
                    "status",
                    "gross_salary",
                    "calculated_at",
                    "finalized_at",
                )
            ),
        )

    def test_tenant_ids_are_not_found_on_pages_and_exports(self):
        category_b = make_category(self.business_b, "Tenant B Category", "B")
        WmsEmployeeCategoryAssignment.objects.create(
            business=self.business_b,
            employee=self.employee_b,
            category=category_b,
        )
        production_services.create_production_entry(
            business=self.business_b,
            location=self.location_b,
            employee=self.employee_b,
            production_date=self.report_date,
            daily_total_pieces=99,
            notes="Tenant B only",
            assignment_quantities={str(self.employee_b.category_assignments.get().public_id): 99},
            user=self.owner_b,
        )
        page_cases = (
            (
                "report_daily_production",
                {
                    "report_date": "2026-07-12",
                    "employee": str(self.employee_b.public_id),
                },
            ),
            (
                "report_daily_production",
                {
                    "report_date": "2026-07-12",
                    "location": str(self.location_b.public_id),
                },
            ),
            (
                "report_daily_production",
                {
                    "report_date": "2026-07-12",
                    "category": str(category_b.public_id),
                },
            ),
            (
                "report_monthly_production",
                {
                    "report_month": "2026-07",
                    "employee": str(self.employee_b.public_id),
                },
            ),
            (
                "report_individual_attendance",
                {
                    "report_month": "2026-07",
                    "employee": str(self.employee_b.public_id),
                },
            ),
            (
                "report_salary",
                {
                    "report_month": "2026-07",
                    "employee": str(self.employee_b.public_id),
                },
            ),
        )
        for name, params in page_cases:
            with self.subTest(name=name, params=params):
                self.assertEqual(
                    self.client.get(reverse(f"wms:{name}"), params).status_code,
                    404,
                )
        self.assertEqual(
            self.client.get(
                reverse("wms:report_daily_production_export"),
                {
                    "report_date": "2026-07-12",
                    "employee": str(self.employee_b.public_id),
                },
            ).status_code,
            404,
        )
        visible = self.client.get(
            reverse("wms:report_daily_production"),
            {"report_date": "2026-07-12"},
        )
        self.assertNotContains(visible, self.employee_b.full_name)
        self.assertEqual(visible.context["report"]["grand_total"], 0)

    def test_location_restriction_applies_to_page_and_export(self):
        location_two_employee = make_employee(
            self.business_a,
            self.location_a2,
            "P8-LOC2",
        )
        category = make_category(self.business_a, "Location Two", "LOC2")
        assignment = WmsEmployeeCategoryAssignment.objects.create(
            business=self.business_a,
            employee=location_two_employee,
            category=category,
        )
        production_services.create_production_entry(
            business=self.business_a,
            location=self.location_a2,
            employee=location_two_employee,
            production_date=self.report_date,
            daily_total_pieces=22,
            notes="Allowed location production",
            assignment_quantities={str(assignment.public_id): 22},
            user=self.owner_a,
        )
        self.create_report_production(daily_total=11)
        user, _access = self.make_staff(
            "phase8-limited-location",
            [
                "wms.reports.view",
                "wms.reports.export",
                "wms.production.view",
            ],
            allowed_locations=(self.location_a2,),
        )
        self.client.force_login(user)
        response = self.client.get(
            reverse("wms:report_daily_production"),
            {"report_date": "2026-07-12"},
        )
        self.assertEqual(response.context["report"]["grand_total"], 22)
        self.assertContains(response, location_two_employee.full_name)
        self.assertNotContains(response, self.piece_employee.full_name)
        export = self.client.get(
            reverse("wms:report_daily_production_export"),
            {"report_date": "2026-07-12"},
        )
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(BytesIO(export.content), read_only=True)
        values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
        self.assertIn(location_two_employee.full_name, values)
        self.assertNotIn(self.piece_employee.full_name, values)

    def test_view_and_export_permissions_are_separate_and_navigation_hides(self):
        viewer, _access = self.make_staff(
            "phase8-viewer",
            ["wms.reports.view", "wms.production.view"],
        )
        self.client.force_login(viewer)
        self.assertEqual(
            self.client.get(reverse("wms:report_index")).status_code,
            200,
        )
        self.assertEqual(
            self.client.get(reverse("wms:report_daily_production_export")).status_code,
            403,
        )
        self.assertContains(
            self.client.get(reverse("wms:report_index")),
            ">Reports</a>",
            html=False,
        )
        self.assertNotContains(
            self.client.get(reverse("wms:report_index")),
            reverse("wms:report_salary"),
        )
        self.assertEqual(
            self.client.get(reverse("wms:report_attendance_summary")).status_code,
            403,
        )
        self.assertEqual(
            self.client.get(reverse("wms:report_salary")).status_code,
            403,
        )

        exporter_only, _access = self.make_staff(
            "phase8-exporter-only",
            ["wms.reports.export", "wms.production.view"],
        )
        self.client.force_login(exporter_only)
        self.assertEqual(
            self.client.get(reverse("wms:report_daily_production_export")).status_code,
            403,
        )

        employee_viewer, _access = self.make_staff(
            "phase8-no-reports",
            ["wms.employees.view"],
        )
        self.client.force_login(employee_viewer)
        employee_page = self.client.get(reverse("wms:employee_list"))
        self.assertEqual(employee_page.status_code, 200)
        self.assertNotContains(
            employee_page,
            reverse("wms:report_index"),
        )
        self.assertEqual(
            self.client.get(reverse("wms:report_index")).status_code,
            403,
        )

    def test_every_export_is_protected_and_excel_contains_context(self):
        self.create_report_production()
        self.create_report_attendance(date(2026, 7, 1))
        self.calculate(self.fixed_employee)
        export_cases = (
            (
                "report_daily_production_export",
                {"report_date": "2026-07-12"},
            ),
            (
                "report_monthly_production_export",
                {
                    "report_month": "2026-07",
                    "employee": str(self.piece_employee.public_id),
                },
            ),
            (
                "report_attendance_summary_export",
                {"date_from": "2026-07-01", "date_to": "2026-07-31"},
            ),
            (
                "report_individual_attendance_export",
                {
                    "report_month": "2026-07",
                    "employee": str(self.fixed_employee.public_id),
                },
            ),
            (
                "report_salary_export",
                {"report_month": "2026-07"},
            ),
        )
        for name, params in export_cases:
            with self.subTest(name=name):
                response = self.client.get(reverse(f"wms:{name}"), params)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response["Content-Type"],
                    ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                )
                workbook = load_workbook(
                    BytesIO(response.content),
                    read_only=True,
                )
                cells = [
                    value for row in workbook.active.iter_rows(values_only=True) for value in row
                ]
                self.assertIn(self.business_a.name, cells)

    def test_excel_neutralizes_formula_injection(self):
        self.create_report_production()
        self.piece_employee.full_name = "=DANGEROUS()"
        self.piece_employee.save()
        response = self.client.get(
            reverse("wms:report_daily_production_export"),
            {"report_date": "2026-07-12"},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        values = [value for row in workbook.active.iter_rows(values_only=True) for value in row]
        self.assertIn("'=DANGEROUS()", values)
        self.assertNotIn("=DANGEROUS()", values)

    def test_representative_selectors_do_not_issue_n_plus_one_queries(self):
        self.create_report_production()
        with CaptureQueriesContext(connection) as captured:
            report = selectors.daily_production(
                self.access_a,
                report_date=self.report_date,
            )
        self.assertEqual(report["grand_total"], 9)
        self.assertLessEqual(len(captured), 3)
