"""Focused coverage for the combined Expenses screen and recurring rows."""
from datetime import date
from decimal import Decimal as D
from io import BytesIO
from unittest.mock import patch
from urllib.parse import urlencode

from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone

from apps.accounts.models import Membership, Role, User
from apps.expenses.models import Expense, RecurringExpenseTemplate
from apps.expenses.services import (
    ensure_recurring_expenses_for_month,
    ensure_recurring_expenses_for_range,
)
from tests.base import TenantTestCase
from tests.test_recurring_expenses import RecurringExpenseTestMixin


class CombinedExpensesPageTests(RecurringExpenseTestMixin, TenantTestCase):
    def setUp(self):
        self.client.force_login(self.owner_a)

    @property
    def fixed_anchor(self):
        return reverse("expenses:list") + "#fixed-expenses"

    def template_payload(self, **overrides):
        next_year = timezone.localdate().year + 1
        values = {
            "name": "Internet",
            "category": str(self.category_a.pk),
            "default_amount": "35.500",
            "due_day": "12",
            "start_date": f"{next_year}-01-01",
            "end_date": "",
            "notes": "Office internet",
            "is_active": "on",
        }
        values.update(overrides)
        return values

    def test_main_page_shows_both_sections_in_required_order(self):
        today = timezone.localdate()
        manual = self.make_manual_expense(
            expense_date=today,
            payee="Daily maintenance",
        )
        template = self.make_template(
            name="Shop Rent",
            start_date=today.replace(day=1),
        )

        response = self.client.get(reverse("expenses:list"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        current_heading = "Current Expenses (Daily Expenses)"
        fixed_heading = "Fixed Expenses (Rent, Phone bills etc)"
        self.assertContains(response, current_heading)
        self.assertContains(response, fixed_heading)
        self.assertLess(content.index(current_heading), content.index(fixed_heading))
        self.assertContains(response, manual.payee)
        self.assertContains(response, template.name)
        self.assertContains(response, "Add Current Expense")
        self.assertContains(response, "Add Fixed Expense")
        self.assertEqual(list(response.context["page_obj"]), [manual])
        self.assertEqual(list(response.context["fixed_templates"]), [template])

    def test_combined_page_uses_simple_labels_only(self):
        self.make_template()
        response = self.client.get(reverse("expenses:list"))

        for expected in (
            "Expense Name / Payee",
            "Monthly Amount",
            "Due Day",
            "Start Date",
            "End Date",
            "Active",
        ):
            self.assertContains(response, expected)
        for hidden_term in (
            "Variable Expenses",
            "Recurring Expense Templates",
            "Provenance",
            "Generated Expense",
            "Generation Month",
        ):
            self.assertNotContains(response, hidden_term)

    def test_main_navigation_points_to_combined_expenses_page(self):
        response = self.client.get(reverse("expenses:list"))
        self.assertContains(
            response,
            f'<a class="nav-link active" href="{reverse("expenses:list")}">',
        )

    def test_legacy_fixed_expense_bookmark_redirects_to_combined_page(self):
        response = self.client.get(reverse("expenses:recurring_list"))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, self.fixed_anchor)

    def test_page_access_never_generates_recurring_expenses(self):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        template = self.make_template(start_date=month_start)

        first = self.client.get(reverse("expenses:list"))
        second = self.client.get(reverse("expenses:list"))

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertFalse(template.generated_expenses.exists())

    def test_page_access_preserves_only_explicitly_generated_months(self):
        today = timezone.localdate()
        template = self.make_template(start_date=today.replace(day=1))
        ensure_recurring_expenses_for_month(self.business_a, today)
        before = list(
            template.generated_expenses.values_list(
                "generated_for_month",
                flat=True,
            )
        )

        response = self.client.get(reverse("expenses:list"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            list(
                template.generated_expenses.values_list(
                    "generated_for_month",
                    flat=True,
                )
            ),
            before,
        )
        self.assertEqual(before, [today.replace(day=1)])

    def test_current_expense_filters_remain_scoped_to_top_section(self):
        matching = self.make_manual_expense(
            category=self.category_a,
            expense_date=date(2026, 7, 10),
            status=Expense.Status.APPROVED,
            payee="Matching current expense",
        )
        self.make_manual_expense(
            category=self.category_a_2,
            expense_date=date(2026, 7, 10),
            payee="Different category",
        )
        self.make_manual_expense(
            category=self.category_a,
            expense_date=date(2026, 6, 30),
            payee="Outside date range",
        )
        self.make_manual_expense(
            category=self.category_a,
            expense_date=date(2026, 7, 11),
            status=Expense.Status.REJECTED,
            payee="Different status",
        )
        fixed = self.make_template(name="Visible fixed definition")
        query = urlencode({
            "category": self.category_a.pk,
            "status": Expense.Status.APPROVED,
            "from": "2026-07-01",
            "to": "2026-07-31",
        })

        response = self.client.get(f"{reverse('expenses:list')}?{query}")

        self.assertEqual(list(response.context["page_obj"]), [matching])
        self.assertEqual(list(response.context["fixed_templates"]), [fixed])

    def test_fixed_create_edit_and_actions_return_to_fixed_section(self):
        create_response = self.client.post(
            reverse("expenses:recurring_create"),
            self.template_payload(),
        )
        self.assertEqual(create_response.url, self.fixed_anchor)
        template = RecurringExpenseTemplate.objects.get(name="Internet")

        edit_response = self.client.post(
            reverse("expenses:recurring_edit", args=[template.public_id]),
            self.template_payload(name="Business Internet"),
        )
        self.assertEqual(edit_response.url, self.fixed_anchor)
        template.refresh_from_db()
        self.assertEqual(template.name, "Business Internet")

        archive_response = self.client.post(
            reverse(
                "expenses:recurring_action",
                args=[template.public_id, "archive"],
            )
        )
        self.assertEqual(archive_response.url, self.fixed_anchor)
        template.refresh_from_db()
        self.assertFalse(template.is_active)

        restore_response = self.client.post(
            reverse(
                "expenses:recurring_action",
                args=[template.public_id, "restore"],
            )
        )
        self.assertEqual(restore_response.url, self.fixed_anchor)
        template.refresh_from_db()
        self.assertTrue(template.is_active)

        delete_response = self.client.post(
            reverse("expenses:recurring_delete", args=[template.public_id])
        )
        self.assertEqual(delete_response.url, self.fixed_anchor)
        self.assertFalse(
            RecurringExpenseTemplate.objects.filter(pk=template.pk).exists()
        )

    def test_view_only_user_sees_sections_without_management_actions(self):
        template = self.make_template()
        viewer = User.objects.create_user(
            email="combined-expense-viewer@example.com",
            password="StrongPass123!",
            full_name="Combined Expense Viewer",
        )
        role = Role.objects.create(
            business=self.business_a,
            name="Combined Expense Viewer",
            permissions=["expenses.view"],
        )
        Membership.objects.create(
            business=self.business_a,
            user=viewer,
            role=role,
        )
        self.client.force_login(viewer)

        response = self.client.get(reverse("expenses:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, template.name)
        self.assertContains(response, "Current Expenses (Daily Expenses)")
        self.assertContains(response, "Fixed Expenses (Rent, Phone bills etc)")
        self.assertNotContains(response, "Add Current Expense")
        self.assertNotContains(response, "Add Fixed Expense")
        self.assertNotContains(
            response,
            reverse("expenses:recurring_edit", args=[template.public_id]),
        )
        self.assertEqual(
            self.client.get(reverse("expenses:recurring_create")).status_code,
            403,
        )

    def test_user_without_expense_view_permission_is_denied(self):
        self.client.force_login(self.cashier_a)
        self.assertEqual(
            self.client.get(reverse("expenses:list")).status_code,
            403,
        )

    def test_combined_page_is_tenant_isolated(self):
        template_a = self.make_template(name="Alpha fixed expense")
        template_b = self.make_template(
            business=self.business_b,
            name="Beta fixed expense",
        )

        response = self.client.get(reverse("expenses:list"))

        self.assertContains(response, template_a.name)
        self.assertNotContains(response, template_b.name)
        self.assertEqual(
            list(response.context["fixed_templates"]),
            [template_a],
        )


class AutomaticFixedExpenseReportTests(
    RecurringExpenseTestMixin,
    TenantTestCase,
):
    def setUp(self):
        self.client.force_login(self.owner_a)

    def report_url(self, start, end, *, key="expenses", export=None, branch=None):
        params = {"from": str(start), "to": str(end)}
        if export:
            params["export"] = export
        if branch:
            params["branch"] = branch
        return f"{reverse('reports:view', args=[key])}?{urlencode(params)}"

    def test_one_month_report_reads_explicit_fixed_expense_without_mutation(self):
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            due_day=12,
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_before = template.generated_expenses.count()
        url = self.report_url(date(2026, 7, 1), date(2026, 7, 31))

        first = self.client.get(url)
        second = self.client.get(url)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(generated_before, 1)
        self.assertEqual(template.generated_expenses.count(), generated_before)
        self.assertEqual(len(first.context["data"]["rows"]), 1)
        self.assertEqual(first.context["data"]["rows"][0][3], "Fixed")

    def test_multi_month_report_reads_explicit_months_without_mutation(self):
        template = self.make_template(
            start_date=date(2026, 1, 1),
            end_date=date(2026, 3, 31),
            default_amount=D("20.000"),
        )
        ensure_recurring_expenses_for_range(
            self.business_a,
            date(2026, 1, 1),
            date(2026, 3, 31),
        )
        generated_before = template.generated_expenses.count()
        url = self.report_url(date(2026, 1, 1), date(2026, 3, 31))

        first = self.client.get(url)
        second = self.client.get(url)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(generated_before, 3)
        self.assertEqual(template.generated_expenses.count(), generated_before)
        self.assertEqual(
            set(
                template.generated_expenses.values_list(
                    "generated_for_month",
                    flat=True,
                )
            ),
            {date(2026, 1, 1), date(2026, 2, 1), date(2026, 3, 1)},
        )
        self.assertEqual(first.context["data"]["totals"][6], D("60.000"))

    def test_report_does_not_generate_unrequested_future_months(self):
        template = self.make_template(start_date=date(2026, 1, 1))
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 1, 1))
        generated_before = list(
            template.generated_expenses.values_list(
                "generated_for_month",
                flat=True,
            )
        )

        response = self.client.get(
            self.report_url(date(2026, 1, 1), date(2026, 1, 31))
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            list(
                template.generated_expenses.values_list(
                    "generated_for_month",
                    flat=True,
                )
            ),
            generated_before,
        )
        self.assertEqual(generated_before, [date(2026, 1, 1)])

    def test_inactive_and_out_of_range_fixed_expenses_are_not_generated(self):
        inactive = self.make_template(
            name="Inactive rent",
            is_active=False,
            start_date=date(2026, 1, 1),
        )
        future = self.make_template(
            name="Future rent",
            start_date=date(2026, 4, 1),
        )
        expense_count_before = Expense.objects.count()

        response = self.client.get(
            self.report_url(date(2026, 1, 1), date(2026, 3, 31))
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(inactive.generated_expenses.exists())
        self.assertFalse(future.generated_expenses.exists())
        self.assertEqual(Expense.objects.count(), expense_count_before)
        self.assertEqual(response.context["data"]["rows"], [])

    def test_report_has_current_and_fixed_rows_with_one_combined_total(self):
        manual = self.make_manual_expense(
            expense_date=date(2026, 7, 10),
            amount=D("5.000"),
        )
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            default_amount=D("250.000"),
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_before = template.generated_expenses.count()

        response = self.client.get(
            self.report_url(date(2026, 7, 1), date(2026, 7, 31))
        )

        data = response.context["data"]
        self.assertEqual({row[3] for row in data["rows"]}, {"Current", "Fixed"})
        self.assertEqual(data["totals"][6], D("255.000"))
        self.assertEqual(len(data["rows"]), 2)
        self.assertEqual(generated_before, 1)
        self.assertEqual(template.generated_expenses.count(), generated_before)
        self.assertIn(manual.expense_number, {row[0] for row in data["rows"]})

    def test_expense_report_date_and_branch_filters_still_apply(self):
        self.make_manual_expense(
            expense_date=date(2026, 7, 10),
            amount=D("5.000"),
        )
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            due_day=20,
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_before = template.generated_expenses.count()

        included = self.client.get(
            self.report_url(
                date(2026, 7, 1),
                date(2026, 7, 31),
                branch=self.branch_a.pk,
            )
        )
        excluded_by_date = self.client.get(
            self.report_url(
                date(2026, 7, 1),
                date(2026, 7, 15),
                branch=self.branch_a.pk,
            )
        )
        excluded_by_branch = self.client.get(
            self.report_url(
                date(2026, 7, 1),
                date(2026, 7, 31),
                branch=self.branch_b.pk,
            )
        )

        self.assertEqual(len(included.context["data"]["rows"]), 2)
        self.assertEqual(len(excluded_by_date.context["data"]["rows"]), 1)
        self.assertEqual(excluded_by_branch.context["data"]["rows"], [])
        self.assertEqual(generated_before, 1)
        self.assertEqual(template.generated_expenses.count(), generated_before)

    def test_csv_xlsx_and_pdf_include_current_and_fixed_rows(self):
        self.make_manual_expense(
            expense_date=date(2026, 7, 10),
            amount=D("5.000"),
        )
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            default_amount=D("250.000"),
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_before = template.generated_expenses.count()
        start, end = date(2026, 7, 1), date(2026, 7, 31)

        csv_response = self.client.get(
            self.report_url(start, end, export="csv")
        )
        csv_body = csv_response.content.decode("utf-8")
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn("Current", csv_body)
        self.assertIn("Fixed", csv_body)
        self.assertIn("255.000", csv_body)

        xlsx_response = self.client.get(
            self.report_url(start, end, export="xlsx")
        )
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
        worksheet = workbook.active
        sources = {worksheet.cell(row=row, column=4).value for row in (2, 3)}
        self.assertEqual(sources, {"Current", "Fixed"})
        self.assertEqual(worksheet.cell(row=4, column=7).value, 255)
        workbook.close()

        pdf_response = HttpResponse(b"%PDF-test", content_type="application/pdf")
        with patch(
            "apps.reports.views.exports.export_pdf",
            return_value=pdf_response,
        ) as export_pdf:
            response = self.client.get(
                self.report_url(start, end, export="pdf")
            )
        self.assertEqual(response.status_code, 200)
        pdf_data = export_pdf.call_args.args[1]
        self.assertEqual({row[3] for row in pdf_data["rows"]}, {"Current", "Fixed"})
        self.assertEqual(pdf_data["totals"][6], D("255.000"))
        self.assertEqual(generated_before, 1)
        self.assertEqual(template.generated_expenses.count(), generated_before)

    def test_other_financial_reports_read_same_row_without_mutation(self):
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            default_amount=D("20.000"),
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_before = template.generated_expenses.count()
        for key in ("profit", "profit_loss", "cash_flow", "expense_analysis"):
            with self.subTest(key=key):
                response = self.client.get(
                    self.report_url(
                        date(2026, 7, 1),
                        date(2026, 7, 31),
                        key=key,
                    )
                )
                self.assertEqual(response.status_code, 200)
                self.assertEqual(generated_before, 1)
                self.assertEqual(
                    template.generated_expenses.count(),
                    generated_before,
                )

    def test_oversized_report_range_is_read_only(self):
        template = self.make_template(start_date=date(2000, 1, 1))
        expense_count_before = Expense.objects.count()

        response = self.client.get(
            self.report_url(date(2000, 1, 1), date(2010, 1, 31))
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(template.generated_expenses.exists())
        self.assertEqual(Expense.objects.count(), expense_count_before)

    def test_report_reads_only_explicit_rows_for_current_tenant(self):
        template_a = self.make_template(
            name="Alpha rent",
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
        )
        template_b = self.make_template(
            business=self.business_b,
            name="Beta rent",
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
        )
        ensure_recurring_expenses_for_month(self.business_a, date(2026, 7, 1))
        generated_a_before = template_a.generated_expenses.count()
        generated_b_before = template_b.generated_expenses.count()

        response = self.client.get(
            self.report_url(date(2026, 7, 1), date(2026, 7, 31))
        )

        self.assertEqual(generated_a_before, 1)
        self.assertEqual(generated_b_before, 0)
        self.assertEqual(template_a.generated_expenses.count(), generated_a_before)
        self.assertEqual(template_b.generated_expenses.count(), generated_b_before)
        numbers = {row[0] for row in response.context["data"]["rows"]}
        self.assertIn(template_a.generated_expenses.get().expense_number, numbers)
        self.assertNotIn(f"REC-202607-{template_b.pk}", numbers)

    def test_report_permission_is_checked_before_generation(self):
        template = self.make_template(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
        )
        viewer = User.objects.create_user(
            email="expenses-only@example.com",
            password="StrongPass123!",
            full_name="Expenses Only",
        )
        role = Role.objects.create(
            business=self.business_a,
            name="Expenses Only",
            permissions=["expenses.view"],
        )
        Membership.objects.create(
            business=self.business_a,
            user=viewer,
            role=role,
        )
        self.client.force_login(viewer)

        response = self.client.get(
            self.report_url(date(2026, 7, 1), date(2026, 7, 31))
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(template.generated_expenses.exists())
