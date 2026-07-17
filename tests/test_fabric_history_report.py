import csv
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.models import Membership, Role, User
from apps.branches.models import Branch, Warehouse
from apps.catalog.models import Brand, Product, ProductVariant, Unit
from apps.inventory import services as inventory
from apps.inventory import workflows as inventory_workflows
from apps.inventory.models import StockMovement
from apps.purchases import services as purchases
from apps.reports.queries import fabric_history
from apps.sales import services as sales
from apps.sales.models import Sale, SaleReturn
from apps.suppliers.models import Supplier

from .base import TenantTestCase

D = Decimal


class FabricHistoryReportTests(TenantTestCase):
    COLUMNS = [
        "Brand",
        "Color",
        "Product / Variant",
        "Opening Stock (Meters)",
        "Purchased (Meters)",
        "Used (Meters)",
        "Remaining (Meters)",
        "Orders Count",
    ]
    DATE_FROM = date(2026, 7, 1)
    DATE_TO = date(2026, 7, 31)

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.meter_a = Unit.objects.for_business(cls.business_a).get(is_meter=True)
        cls.piece_a = Unit.objects.for_business(cls.business_a).get(name="Piece")
        cls.brand = Brand.objects.create(
            business=cls.business_a,
            name="Shared Loom",
        )
        cls.empty_brand = Brand.objects.create(
            business=cls.business_a,
            name="Empty Brand",
        )

        cls.fabric_one = Product.objects.create(
            business=cls.business_a,
            name="Royal One",
            sku="FAB-ROYAL-ONE",
            brand=cls.brand,
            unit=cls.meter_a,
            product_type=Product.Type.VARIANT,
            track_inventory=True,
            is_tailoring_item=True,
        )
        cls.black_one = ProductVariant.objects.create(
            business=cls.business_a,
            product=cls.fabric_one,
            name="Black",
            attributes={"Color": "Black"},
            sku="FAB-ROYAL-ONE-BLACK",
            purchase_price=D("2.000"),
        )
        cls.fabric_two = Product.objects.create(
            business=cls.business_a,
            name="Royal Two",
            sku="FAB-ROYAL-TWO",
            brand=cls.brand,
            unit=cls.meter_a,
            product_type=Product.Type.VARIANT,
            track_inventory=True,
            is_tailoring_item=True,
        )
        cls.black_two = ProductVariant.objects.create(
            business=cls.business_a,
            product=cls.fabric_two,
            name="Black",
            attributes={},
            sku="FAB-ROYAL-TWO-BLACK",
            purchase_price=D("3.000"),
        )
        cls.unbranded = Product.objects.create(
            business=cls.business_a,
            name="Ivory",
            sku="FAB-IVORY",
            unit=cls.meter_a,
            product_type=Product.Type.STANDARD,
            track_inventory=True,
            is_tailoring_item=True,
            purchase_price=D("1.500"),
        )

        cls.second_branch = Branch.objects.create(
            business=cls.business_a,
            name="Seeb",
            code="SEEB-FAB",
        )
        cls.second_warehouse = Warehouse.objects.create(
            business=cls.business_a,
            branch=cls.second_branch,
            name="Seeb Fabric",
            code="SEEB-FAB",
        )
        for warehouse, product, variant, quantity, cost in (
            (
                cls.warehouse_a,
                cls.fabric_one,
                cls.black_one,
                D("20.000"),
                D("2.000"),
            ),
            (
                cls.second_warehouse,
                cls.fabric_one,
                cls.black_one,
                D("7.000"),
                D("2.000"),
            ),
            (
                cls.warehouse_a,
                cls.fabric_two,
                cls.black_two,
                D("30.000"),
                D("3.000"),
            ),
            (
                cls.warehouse_a,
                cls.unbranded,
                None,
                D("5.000"),
                D("1.500"),
            ),
        ):
            inventory.set_opening_stock(
                business=cls.business_a,
                warehouse=warehouse,
                product=product,
                variant=variant,
                quantity=quantity,
                unit_cost=cost,
                user=cls.owner_a,
            )

        cls.retail = Product.objects.create(
            business=cls.business_a,
            name="Finished Kumma",
            sku="RETAIL-KUMMA",
            unit=cls.piece_a,
            product_type=Product.Type.STANDARD,
            track_inventory=True,
            is_tailoring_item=True,
            sale_price=D("10.000"),
        )
        inventory.set_opening_stock(
            business=cls.business_a,
            warehouse=cls.warehouse_a,
            product=cls.retail,
            quantity=D("10.000"),
            unit_cost=D("1.000"),
            user=cls.owner_a,
        )
        cls.legacy_tailoring = Product.objects.create(
            business=cls.business_a,
            name="Legacy Estimate Garment",
            sku="LEGACY-ESTIMATE",
            unit=None,
            product_type=Product.Type.STANDARD,
            track_inventory=True,
            is_tailoring_item=True,
            estimated_adult_fabric=D("3.500"),
            estimated_child_fabric=D("2.250"),
            sale_price=D("10.000"),
        )
        inventory.set_opening_stock(
            business=cls.business_a,
            warehouse=cls.warehouse_a,
            product=cls.legacy_tailoring,
            quantity=D("10.000"),
            unit_cost=D("1.000"),
            user=cls.owner_a,
        )

        cls.supplier = Supplier.objects.create(
            business=cls.business_a,
            code="FAB-SUP",
            name="Fabric Supplier",
        )

        cls.meter_b = Unit.objects.for_business(cls.business_b).get(is_meter=True)
        cls.brand_b = Brand.objects.create(
            business=cls.business_b,
            name="Tenant B Loom",
        )
        cls.fabric_b = Product.objects.create(
            business=cls.business_b,
            name="Tenant B Fabric",
            sku="TENANT-B-FAB",
            brand=cls.brand_b,
            unit=cls.meter_b,
            product_type=Product.Type.VARIANT,
            track_inventory=True,
            is_tailoring_item=True,
        )
        cls.variant_b = ProductVariant.objects.create(
            business=cls.business_b,
            product=cls.fabric_b,
            name="Black",
            attributes={"Color": "Black"},
            sku="TENANT-B-FAB-BLACK",
        )
        inventory.set_opening_stock(
            business=cls.business_b,
            warehouse=cls.warehouse_b,
            product=cls.fabric_b,
            variant=cls.variant_b,
            quantity=D("99.000"),
            unit_cost=D("1.000"),
            user=cls.owner_b,
        )

    def setUp(self):
        self.allow_no_shift()
        self.client.force_login(self.owner_a)

    def report(self, **overrides):
        filters = {
            "date_from": self.DATE_FROM,
            "date_to": self.DATE_TO,
        }
        filters.update(overrides)
        data = fabric_history(self.business_a, filters)
        self.assertEqual(data["columns"], self.COLUMNS)
        return data

    def detail_rows(self, data):
        return [
            dict(zip(data["columns"], row, strict=True))
            for row in data["rows"]
            if not str(row[0]).startswith("Brand Total - ")
        ]

    def detail(self, data, item_name):
        matches = [row for row in self.detail_rows(data) if row["Product / Variant"] == item_name]
        self.assertEqual(len(matches), 1, (item_name, self.detail_rows(data)))
        return matches[0]

    def brand_total(self, data, brand_name):
        matches = [row for row in data["brand_totals"] if row["brand"] == brand_name]
        self.assertEqual(len(matches), 1, data["brand_totals"])
        return matches[0]

    def meter_sale(
        self,
        *,
        product=None,
        variant=None,
        meters="3.500",
        warehouse=None,
        extra_lines=None,
    ):
        product = product or self.fabric_one
        variant = self.black_one if variant is None and product == self.fabric_one else variant
        warehouse = warehouse or self.warehouse_a
        branch = warehouse.branch or self.branch_a
        lines = [
            {
                "product": product,
                "variant": variant,
                "quantity": D("1"),
                "unit_price": D("25.000"),
                "fabric_meter_used": str(meters),
                "garment_classification": "adult",
                "collection_type": "normal",
                "tailoring_details": {},
            }
        ]
        lines.extend(extra_lines or [])
        return sales.complete_sale(
            business=self.business_a,
            branch=branch,
            warehouse=warehouse,
            cashier=self.owner_a,
            customer=self.walk_in_a,
            items=lines,
            payments=[
                {
                    "method": self.cash_a,
                    "amount": D("25.000") * len(lines),
                }
            ],
            membership=self.membership_a(),
            delivery_date=self.DATE_TO,
        )

    def purchase_receipt(
        self,
        quantity,
        *,
        warehouse=None,
        purchase_date=None,
        product=None,
        variant=None,
    ):
        warehouse = warehouse or self.warehouse_a
        product = product or self.fabric_one
        variant = self.black_one if variant is None and product == self.fabric_one else variant
        purchase = purchases.create_purchase(
            business=self.business_a,
            supplier=self.supplier,
            branch=warehouse.branch or self.branch_a,
            warehouse=warehouse,
            rows=[
                {
                    "product": product,
                    "variant": variant,
                    "quantity": D(str(quantity)),
                    "unit_cost": D("2.000"),
                }
            ],
            user=self.owner_a,
            purchase_date=purchase_date or self.DATE_FROM,
        )
        item = purchase.items.get()
        purchases.receive_purchase(
            purchase=purchase,
            quantities={item.pk: D(str(quantity))},
            user=self.owner_a,
        )
        return purchase, item

    def set_movement_time(self, *, reference_type, reference_id, posted_at):
        updated = (
            StockMovement.objects.for_business(self.business_a)
            .filter(
                reference_type=reference_type,
                reference_id=reference_id,
            )
            .update(created_at=posted_at)
        )
        self.assertGreater(updated, 0)

    def report_url(self, **params):
        from urllib.parse import urlencode

        base = reverse("reports:view", args=["fabric_history"])
        return f"{base}?{urlencode(params)}" if params else base

    def test_authoritative_product_variant_key_color_fallback_and_totals(self):
        data = self.report()
        one = self.detail(data, "Royal One - Black")
        two = self.detail(data, "Royal Two - Black")
        ivory = self.detail(data, "Ivory")

        self.assertEqual((one["Brand"], one["Color"]), ("Shared Loom", "Black"))
        self.assertEqual((two["Brand"], two["Color"]), ("Shared Loom", "Black"))
        self.assertEqual(one["Opening Stock (Meters)"], D("27.000"))
        self.assertEqual(two["Opening Stock (Meters)"], D("30.000"))
        self.assertEqual(one["Remaining (Meters)"], D("27.000"))
        self.assertEqual(two["Remaining (Meters)"], D("30.000"))
        self.assertEqual((ivory["Brand"], ivory["Color"]), ("No Brand", "Ivory"))
        self.assertEqual(ivory["Opening Stock (Meters)"], D("5.000"))
        self.assertEqual(ivory["Remaining (Meters)"], D("5.000"))

        shared = self.brand_total(data, "Shared Loom")
        self.assertEqual(shared["opening"], D("57.000"))
        self.assertEqual(shared["remaining"], D("57.000"))
        self.assertEqual(shared["purchased"], D("0.000"))
        self.assertEqual(shared["used"], D("0.000"))
        self.assertEqual(shared["orders"], 0)
        self.assertIn(
            [
                "Brand Total - Shared Loom",
                "",
                "",
                D("57.000"),
                D("0.000"),
                D("0.000"),
                D("57.000"),
                0,
            ],
            data["rows"],
        )
        self.assertEqual(
            data["totals"],
            [
                "GRAND TOTAL",
                "",
                "",
                D("62.000"),
                D("0.000"),
                D("0.000"),
                D("62.000"),
                0,
            ],
        )

    def test_opening_stock_is_separate_from_purchases_and_warehouse_scoped(self):
        before = self.detail(self.report(), "Royal One - Black")
        self.assertEqual(before["Opening Stock (Meters)"], D("27.000"))
        self.assertEqual(before["Purchased (Meters)"], D("0.000"))

        purchase, _item = self.purchase_receipt("5.000")
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=purchase.purchase_number,
            posted_at=datetime(2026, 7, 8, 8, tzinfo=UTC),
        )

        consolidated = self.detail(self.report(), "Royal One - Black")
        main = self.detail(
            self.report(warehouse_id=self.warehouse_a.id),
            "Royal One - Black",
        )
        second = self.detail(
            self.report(warehouse_id=self.second_warehouse.id),
            "Royal One - Black",
        )
        self.assertEqual(consolidated["Opening Stock (Meters)"], D("27.000"))
        self.assertEqual(consolidated["Purchased (Meters)"], D("5.000"))
        self.assertEqual(main["Opening Stock (Meters)"], D("20.000"))
        self.assertEqual(main["Purchased (Meters)"], D("5.000"))
        self.assertEqual(second["Opening Stock (Meters)"], D("7.000"))
        self.assertEqual(second["Purchased (Meters)"], D("0.000"))

    def test_commercial_opening_used_remaining_example(self):
        fabric = Product.objects.create(
            business=self.business_a,
            name="Commercial Example",
            sku="FAB-COMMERCIAL-EXAMPLE",
            brand=self.brand,
            unit=self.meter_a,
            product_type=Product.Type.VARIANT,
            track_inventory=True,
            is_tailoring_item=True,
        )
        color = ProductVariant.objects.create(
            business=self.business_a,
            product=fabric,
            name="Navy",
            attributes={"Color": "Navy"},
            sku="FAB-COMMERCIAL-EXAMPLE-NAVY",
        )
        inventory.set_opening_stock(
            business=self.business_a,
            warehouse=self.warehouse_a,
            product=fabric,
            variant=color,
            quantity=D("25.000"),
            unit_cost=D("2.000"),
            user=self.owner_a,
        )
        self.meter_sale(
            product=fabric,
            variant=color,
            meters="7.000",
        )

        row = self.detail(self.report(), "Commercial Example - Navy")
        self.assertEqual(row["Opening Stock (Meters)"], D("25.000"))
        self.assertEqual(row["Purchased (Meters)"], D("0.000"))
        self.assertEqual(row["Used (Meters)"], D("7.000"))
        self.assertEqual(row["Remaining (Meters)"], D("18.000"))
        self.assertEqual(row["Orders Count"], 1)

    def test_purchased_uses_posting_period_and_nets_purchase_returns(self):
        purchase, item = self.purchase_receipt(
            "12.000",
            purchase_date=date(2026, 6, 1),
        )
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=purchase.purchase_number,
            posted_at=datetime(2026, 7, 10, 8, tzinfo=UTC),
        )
        purchase_return = purchases.return_purchase(
            purchase=purchase,
            quantities={item.pk: D("3.000")},
            user=self.owner_a,
            reason="Supplier return",
        )
        self.set_movement_time(
            reference_type="PurchaseReturn",
            reference_id=purchase_return.return_number,
            posted_at=datetime(2026, 7, 12, 8, tzinfo=UTC),
        )

        outside, _item = self.purchase_receipt(
            "5.000",
            purchase_date=date(2026, 7, 15),
        )
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=outside.purchase_number,
            posted_at=datetime(2026, 6, 30, 8, tzinfo=UTC),
        )

        row = self.detail(self.report(), "Royal One - Black")
        self.assertEqual(row["Opening Stock (Meters)"], D("27.000"))
        self.assertEqual(row["Purchased (Meters)"], D("9.000"))
        self.assertEqual(row["Remaining (Meters)"], D("41.000"))

    def test_used_comes_only_from_pos_meter_and_excludes_legacy_and_retail(self):
        sale = self.meter_sale(meters="3.125")
        sale.items.update(
            estimated_fabric=D("88.000"),
            actual_fabric_used=D("99.000"),
        )

        sales.complete_sale(
            business=self.business_a,
            branch=self.branch_a,
            warehouse=self.warehouse_a,
            cashier=self.owner_a,
            customer=self.walk_in_a,
            items=[
                {
                    "product": self.retail,
                    "quantity": D("2"),
                    "unit_price": D("10.000"),
                    "fabric_meter_used": "7.500",
                }
            ],
            payments=[{"method": self.cash_a, "amount": D("20.000")}],
            membership=self.membership_a(),
        )
        legacy_sale = sales.complete_sale(
            business=self.business_a,
            branch=self.branch_a,
            warehouse=self.warehouse_a,
            cashier=self.owner_a,
            customer=self.walk_in_a,
            items=[
                {
                    "product": self.legacy_tailoring,
                    "quantity": D("1"),
                    "unit_price": D("10.000"),
                    "garment_classification": "adult",
                }
            ],
            payments=[{"method": self.cash_a, "amount": D("10.000")}],
            membership=self.membership_a(),
            delivery_date=self.DATE_TO,
        )
        sales.update_actual_fabric(
            sale_item=legacy_sale.items.get(),
            actual_fabric_used="6.750",
            user=self.owner_a,
            membership=self.membership_a(),
        )

        data = self.report()
        row = self.detail(data, "Royal One - Black")
        self.assertEqual(row["Used (Meters)"], D("3.125"))
        self.assertEqual(row["Orders Count"], 1)
        names = {item["Product / Variant"] for item in self.detail_rows(data)}
        self.assertNotIn(self.retail.name, names)
        self.assertNotIn(self.legacy_tailoring.name, names)

    def test_each_independent_garment_line_counts_as_one_order(self):
        second_line = {
            "product": self.fabric_one,
            "variant": self.black_one,
            "quantity": D("1"),
            "unit_price": D("25.000"),
            "fabric_meter_used": "2.250",
            "garment_classification": "child",
            "collection_type": "premium",
            "tailoring_details": {},
        }
        sale = self.meter_sale(
            meters="3.500",
            extra_lines=[second_line],
        )
        self.assertEqual(sale.items.count(), 2)

        row = self.detail(self.report(), "Royal One - Black")
        self.assertEqual(row["Used (Meters)"], D("5.750"))
        self.assertEqual(row["Orders Count"], 2)

    def test_void_restock_non_restock_and_delivery_cancel_follow_inventory(self):
        voided = self.meter_sale(meters="4.000")
        sales.void_sale(
            sale=voided,
            user=self.owner_a,
            reason="Booking error",
        )

        restocked = self.meter_sale(meters="3.000")
        restocked_item = restocked.items.get()
        sales.process_return(
            sale=restocked,
            items=[
                {
                    "sale_item": restocked_item,
                    "quantity": D("1"),
                    "restock": True,
                }
            ],
            refund_method=SaleReturn.RefundMethod.CASH,
            user=self.owner_a,
        )

        retained = self.meter_sale(meters="2.000")
        retained_item = retained.items.get()
        sales.process_return(
            sale=retained,
            items=[
                {
                    "sale_item": retained_item,
                    "quantity": D("1"),
                    "restock": False,
                }
            ],
            refund_method=SaleReturn.RefundMethod.CASH,
            user=self.owner_a,
            restock=False,
        )

        delivery_cancelled = self.meter_sale(meters="1.500")
        sales.set_delivery_status(
            sale=delivery_cancelled,
            status=Sale.DeliveryStatus.CANCELLED,
            user=self.owner_a,
        )
        restocked.refresh_from_db()
        retained.refresh_from_db()
        delivery_cancelled.refresh_from_db()
        self.assertEqual(restocked.status, Sale.Status.RETURNED)
        self.assertEqual(retained.status, Sale.Status.RETURNED)
        self.assertEqual(delivery_cancelled.status, Sale.Status.COMPLETED)

        row = self.detail(self.report(), "Royal One - Black")
        self.assertEqual(row["Used (Meters)"], D("3.500"))
        self.assertEqual(row["Orders Count"], 2)
        self.assertEqual(row["Remaining (Meters)"], D("23.500"))

    def test_remaining_uses_stocklevel_for_opening_adjustments_and_transfers(self):
        inventory_workflows.create_adjustment(
            business=self.business_a,
            warehouse=self.warehouse_a,
            reason="other",
            rows=[
                {
                    "product": self.fabric_one,
                    "variant": self.black_one,
                    "quantity": D("2.000"),
                }
            ],
            user=self.owner_a,
        )
        transfer = inventory_workflows.create_transfer(
            business=self.business_a,
            from_warehouse=self.warehouse_a,
            to_warehouse=self.second_warehouse,
            rows=[
                {
                    "product": self.fabric_one,
                    "variant": self.black_one,
                    "quantity": D("5.000"),
                }
            ],
            user=self.owner_a,
        )
        inventory_workflows.dispatch_transfer(transfer=transfer, user=self.owner_a)
        inventory_workflows.receive_transfer(transfer=transfer, user=self.owner_a)

        consolidated = self.detail(self.report(), "Royal One - Black")
        main = self.detail(
            self.report(warehouse_id=self.warehouse_a.id),
            "Royal One - Black",
        )
        second = self.detail(
            self.report(warehouse_id=self.second_warehouse.id),
            "Royal One - Black",
        )
        self.assertEqual(consolidated["Remaining (Meters)"], D("29.000"))
        self.assertEqual(main["Remaining (Meters)"], D("17.000"))
        self.assertEqual(second["Remaining (Meters)"], D("12.000"))
        self.assertEqual(consolidated["Opening Stock (Meters)"], D("27.000"))
        self.assertEqual(main["Opening Stock (Meters)"], D("20.000"))
        self.assertEqual(second["Opening Stock (Meters)"], D("7.000"))

    def test_brand_and_warehouse_filters_apply_to_one_shared_dataset(self):
        main_purchase, _item = self.purchase_receipt("4.000")
        second_purchase, _item = self.purchase_receipt(
            "6.000",
            warehouse=self.second_warehouse,
        )
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=main_purchase.purchase_number,
            posted_at=datetime(2026, 7, 5, 8, tzinfo=UTC),
        )
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=second_purchase.purchase_number,
            posted_at=datetime(2026, 7, 6, 8, tzinfo=UTC),
        )
        self.meter_sale(meters="1.000")
        self.meter_sale(meters="2.000", warehouse=self.second_warehouse)

        data = self.report(
            brand_id=self.brand.id,
            warehouse_id=self.second_warehouse.id,
        )
        self.assertEqual(data["detail_count"], 1)
        row = self.detail(data, "Royal One - Black")
        self.assertEqual(row["Opening Stock (Meters)"], D("7.000"))
        self.assertEqual(row["Purchased (Meters)"], D("6.000"))
        self.assertEqual(row["Used (Meters)"], D("2.000"))
        self.assertEqual(row["Remaining (Meters)"], D("11.000"))
        self.assertEqual(row["Orders Count"], 1)
        self.assertEqual(
            data["totals"],
            [
                "GRAND TOTAL",
                "",
                "",
                D("7.000"),
                D("6.000"),
                D("2.000"),
                D("11.000"),
                1,
            ],
        )

    def test_business_timezone_controls_sale_date_and_page_defaults(self):
        self.business_a.timezone = "Asia/Muscat"
        self.business_a.save(update_fields=["timezone", "updated_at"])
        local_july = self.meter_sale(meters="1.250")
        local_june = self.meter_sale(meters="2.500")
        included_at = datetime(2026, 6, 30, 20, 30, tzinfo=UTC)
        excluded_at = datetime(2026, 6, 30, 19, 30, tzinfo=UTC)
        Sale.objects.filter(pk=local_july.pk).update(sale_date=included_at)
        Sale.objects.filter(pk=local_june.pk).update(sale_date=excluded_at)

        row = self.detail(
            self.report(
                date_from=date(2026, 7, 1),
                date_to=date(2026, 7, 1),
            ),
            "Royal One - Black",
        )
        self.assertEqual(row["Used (Meters)"], D("1.250"))
        self.assertEqual(row["Orders Count"], 1)

        with patch("apps.core.date_ranges.timezone.now", return_value=included_at):
            response = self.client.get(self.report_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["filters"]["date_from"], "2026-07-01")
        self.assertEqual(response.context["filters"]["date_to"], "2026-07-01")

    def test_tenant_and_injected_warehouse_filters_cannot_leak_data(self):
        data = self.report()
        names = {row["Product / Variant"] for row in self.detail_rows(data)}
        self.assertNotIn("Tenant B Fabric - Black", names)
        self.assertEqual(data["totals"][3], D("62.000"))
        self.assertEqual(
            self.report(brand_id=self.brand_b.id)["detail_count"],
            0,
        )
        self.assertEqual(
            self.report(warehouse_id=self.warehouse_b.id)["detail_count"],
            0,
        )
        restricted = self.report(
            allowed_branch_ids=[self.branch_a.id],
            warehouse_id=self.second_warehouse.id,
        )
        self.assertEqual(restricted["detail_count"], 0)
        self.assertIsNone(restricted["totals"])

        user = User.objects.create_user(
            email="fabric-branch-viewer@example.com",
            password="StrongPass123!",
            full_name="Fabric Branch Viewer",
        )
        role = Role.objects.create(
            business=self.business_a,
            name="Fabric Branch Viewer",
            permissions=["reports.view"],
        )
        membership = Membership.objects.create(
            business=self.business_a,
            user=user,
            role=role,
        )
        membership.branches.add(self.branch_a)
        self.client.force_login(user)
        response = self.client.get(self.report_url(warehouse=self.second_warehouse.id))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["data"]["detail_count"], 0)
        self.assertNotIn(
            self.second_warehouse,
            list(response.context["warehouses"]),
        )

    def test_page_navigation_empty_state_and_permissions(self):
        index = self.client.get(reverse("reports:index"))
        inventory_group = next(
            group for group in index.context["groups"] if group["name"] == "Inventory"
        )
        report_item = next(
            item for item in inventory_group["items"] if item["key"] == "fabric_history"
        )
        self.assertEqual(report_item["title"], "Fabric History Report")

        page = self.client.get(self.report_url())
        self.assertEqual(page.status_code, 200)
        self.assertEqual(page.context["title"], "Fabric History Report")
        self.assertContains(page, "Fabric History Report")
        self.assertContains(page, 'name="from"')
        self.assertContains(page, 'name="to"')
        self.assertContains(page, 'name="brand"')
        self.assertContains(page, 'name="warehouse"')
        self.assertContains(page, "Opening Stock (Meters)")
        self.assertContains(page, "Print")
        self.assertTrue(page.context["back_enabled"])
        self.assertEqual(page.context["back_url"], reverse("reports:index"))

        empty = self.client.get(self.report_url(brand=self.empty_brand.id))
        self.assertEqual(empty.status_code, 200)
        self.assertEqual(empty.context["data"]["detail_count"], 0)
        self.assertContains(empty, "No data for the selected filters.")

        self.client.force_login(self.cashier_a)
        self.assertEqual(self.client.get(self.report_url()).status_code, 403)
        cashier_index = self.client.get(reverse("reports:index"))
        self.assertEqual(cashier_index.status_code, 403)

    def test_csv_xlsx_and_pdf_use_the_screen_dataset_and_meter_precision(self):
        purchase, _item = self.purchase_receipt("5.000")
        self.set_movement_time(
            reference_type="Purchase",
            reference_id=purchase.purchase_number,
            posted_at=datetime(2026, 7, 8, 8, tzinfo=UTC),
        )
        self.meter_sale(meters="1.500")
        params = {
            "from": self.DATE_FROM.isoformat(),
            "to": self.DATE_TO.isoformat(),
            "brand": self.brand.id,
            "warehouse": self.warehouse_a.id,
        }
        screen = self.client.get(self.report_url(**params))
        self.assertEqual(screen.status_code, 200)
        screen_data = screen.context["data"]

        csv_response = self.client.get(self.report_url(**params, export="csv"))
        self.assertEqual(csv_response.status_code, 200)
        csv_rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8-sig"))))
        self.assertEqual(csv_rows[0], self.COLUMNS)
        expected_rows = [
            ["-" if value in (None, "") else str(value) for value in row]
            for row in screen_data["rows"]
        ]
        self.assertEqual(csv_rows[1 : 1 + len(expected_rows)], expected_rows)
        grand_index = 1 + len(expected_rows)
        self.assertEqual(
            csv_rows[grand_index],
            ["-" if value in (None, "") else str(value) for value in screen_data["totals"]],
        )

        xlsx_response = self.client.get(self.report_url(**params, export="xlsx"))
        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True)
        sheet = workbook.active
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            self.COLUMNS,
        )
        first_detail = self.detail(screen_data, "Royal One - Black")
        xlsx_detail_row = next(
            row for row in sheet.iter_rows(min_row=2) if row[2].value == "Royal One - Black"
        )
        for column_index, key in (
            (3, "Opening Stock (Meters)"),
            (4, "Purchased (Meters)"),
            (5, "Used (Meters)"),
            (6, "Remaining (Meters)"),
        ):
            cell = xlsx_detail_row[column_index]
            self.assertEqual(D(str(cell.value)), first_detail[key])
            self.assertEqual(cell.number_format, "0.000")

        with patch(
            "apps.reports.exports.render_pdf",
            return_value=b"%PDF-1.4 fabric-history",
        ) as render_pdf:
            pdf_response = self.client.get(self.report_url(**params, export="pdf"))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        pdf_context = render_pdf.call_args.args[1]
        self.assertEqual(pdf_context["data"], screen_data)
        self.assertIn("2026-07-01", pdf_context["filters_label"])
        self.assertIn("2026-07-31", pdf_context["filters_label"])
