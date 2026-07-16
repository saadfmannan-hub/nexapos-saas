"""Supplier Payments & Cheques report and export coverage."""
import csv
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.template.loader import render_to_string
from django.urls import reverse

from apps.accounts.models import Membership, Role, User
from apps.branches.models import Branch, Warehouse
from apps.core.date_ranges import business_localdate
from apps.purchases import services as purchases
from apps.reports.exports import _cell
from apps.reports.queries import REPORTS, supplier_balances
from apps.suppliers.models import Supplier, SupplierPayment

from .base import TenantTestCase

D = Decimal


class SupplierPaymentsChequeReportTests(TenantTestCase):
    key = "supplier_payments_cheques"
    columns = [
        "Date", "Supplier", "Purchase No.", "Purchase Date", "Pmt Medium",
        "Amount", "Cheque Number", "Bank Name", "Cheque Issue Date",
        "Cheque Payment Date", "Cheque Status", "Paid", "Cheques Pending",
        "Remaining Balance", "Supplier Balance",
    ]

    def setUp(self):
        self.business_a.timezone = "Asia/Muscat"
        self.business_a.save(update_fields=["timezone", "updated_at"])
        self.today = business_localdate(self.business_a)
        self.issue_date = self.today - timedelta(days=5)
        self.payment_date = self.today + timedelta(days=15)
        self.supplier = Supplier.objects.create(
            business=self.business_a, code="RPT-SUP-1", name="Report Supplier One",
        )
        self.other_supplier = Supplier.objects.create(
            business=self.business_a, code="RPT-SUP-2", name="Report Supplier Two",
        )
        self.other_branch = Branch.objects.create(
            business=self.business_a, name="Report Branch", code="RPT-BR",
        )
        self.other_warehouse = Warehouse.objects.create(
            business=self.business_a,
            branch=self.other_branch,
            name="Report Warehouse",
            code="RPT-WH",
        )
        self.purchase = self.make_purchase(
            supplier=self.supplier,
            branch=self.branch_a,
            warehouse=self.warehouse_a,
            quantity="100",
        )
        self.other_purchase = self.make_purchase(
            supplier=self.other_supplier,
            branch=self.other_branch,
            warehouse=self.other_warehouse,
            quantity="50",
        )

        created = purchases.record_purchase_payments(
            purchase=self.purchase,
            rows=[
                self.immediate(SupplierPayment.Method.CASH, "100"),
                self.immediate(SupplierPayment.Method.BANK, "75"),
                self.immediate(SupplierPayment.Method.CARD, "50"),
                self.cheque("100", "CHQ-PENDING", "Bank Muscat"),
                self.cheque("125", "CHQ-CLEARED", "Sohar International"),
                self.cheque("80", "CHQ-BOUNCED", "National Bank of Oman"),
                self.cheque("70", "CHQ-CANCELLED", "Bank Dhofar"),
            ],
            user=self.owner_a,
        )
        self.payments = {payment.cheque_number or payment.method: payment for payment in created}
        purchases.update_cheque_status(
            payment=self.payments["CHQ-CLEARED"],
            status=SupplierPayment.ChequeStatus.CLEARED,
            user=self.owner_a,
        )
        purchases.update_cheque_status(
            payment=self.payments["CHQ-BOUNCED"],
            status=SupplierPayment.ChequeStatus.BOUNCED,
            user=self.owner_a,
        )
        purchases.update_cheque_status(
            payment=self.payments["CHQ-CANCELLED"],
            status=SupplierPayment.ChequeStatus.CANCELLED,
            user=self.owner_a,
        )
        self.other_payment = purchases.record_purchase_payments(
            purchase=self.other_purchase,
            rows=[self.immediate(SupplierPayment.Method.CASH, "40")],
            user=self.owner_a,
        )[0]
        self.purchase.refresh_from_db()
        self.other_purchase.refresh_from_db()
        self.supplier.refresh_from_db()
        self.other_supplier.refresh_from_db()
        self.client.force_login(self.owner_a)

    def make_purchase(self, *, supplier, branch, warehouse, quantity):
        purchase = purchases.create_purchase(
            business=self.business_a,
            supplier=supplier,
            branch=branch,
            warehouse=warehouse,
            rows=[{
                "product": self.product_a,
                "variant": None,
                "quantity": D(quantity),
                "unit_cost": D("10.000"),
            }],
            user=self.owner_a,
            purchase_date=self.today,
        )
        item = purchase.items.get()
        purchases.receive_purchase(
            purchase=purchase,
            quantities={item.pk: item.quantity_ordered},
            user=self.owner_a,
        )
        purchase.refresh_from_db()
        return purchase

    def immediate(self, method, amount):
        return {"method": method, "amount": D(amount)}

    def cheque(self, amount, number, bank):
        return {
            "method": SupplierPayment.Method.CHEQUE,
            "amount": D(amount),
            "cheque_number": number,
            "bank_name": bank,
            "cheque_issue_date": self.issue_date,
            "due_date": self.payment_date,
        }

    def report_url(self, **params):
        url = reverse("reports:view", args=[self.key])
        if not params:
            return url
        from urllib.parse import urlencode

        return f"{url}?{urlencode(params)}"

    def report_data(self, **params):
        response = self.client.get(self.report_url(**params))
        self.assertEqual(response.status_code, 200)
        return response.context["data"]

    def test_report_center_lists_report_third_under_purchasing(self):
        response = self.client.get(reverse("reports:index"))
        purchasing = next(group for group in response.context["groups"] if group["name"] == "Purchasing")
        self.assertEqual(
            [item["title"] for item in purchasing["items"]],
            ["Purchases", "Outstanding supplier balances", "Supplier Payments & Cheques"],
        )
        self.assertEqual(purchasing["items"][2]["key"], self.key)
        report = self.client.get(self.report_url())
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report.context["title"], "Supplier Payments & Cheques")

    def test_report_requires_financial_and_export_permissions(self):
        user = User.objects.create_user(
            email="payment-report-viewer@example.com",
            password="StrongPass123!",
            full_name="Payment Report Viewer",
        )
        role = Role.objects.create(
            business=self.business_a,
            name="Payment Report Viewer",
            permissions=["reports.view"],
        )
        Membership.objects.create(business=self.business_a, user=user, role=role)
        self.client.force_login(user)
        self.assertEqual(self.client.get(self.report_url()).status_code, 403)
        index = self.client.get(reverse("reports:index"))
        self.assertNotContains(index, "Supplier Payments &amp; Cheques")

        role.permissions = ["reports.view", "reports.financial"]
        role.save(update_fields=["permissions", "updated_at"])
        self.client.force_login(user)
        self.assertEqual(self.client.get(self.report_url()).status_code, 200)
        self.assertEqual(
            self.client.get(self.report_url(export="csv")).status_code,
            403,
        )

    def test_exact_columns_and_one_row_per_payment_record(self):
        data = self.report_data()
        self.assertEqual(data["columns"], self.columns)
        self.assertEqual(len(data["rows"]), 8)
        self.assertEqual(
            len(data["rows"]),
            SupplierPayment.objects.for_business(self.business_a).count(),
        )

    def test_cash_bank_and_card_rows_have_blank_cheque_fields(self):
        rows = self.report_data()["rows"]
        expected = {
            D("100.000"): "Cash",
            D("75.000"): "Bank Transfer",
            D("50.000"): "Card",
        }
        for amount, label in expected.items():
            with self.subTest(label=label):
                row = next(
                    item for item in rows
                    if item[5] == amount and item[4] == label and item[6] is None
                )
                self.assertEqual(row[6:11], [None, None, None, None, None])

    def test_cheque_rows_show_details_due_date_and_all_statuses(self):
        rows = {row[6]: row for row in self.report_data()["rows"] if row[6]}
        self.assertEqual(set(rows), {
            "CHQ-PENDING", "CHQ-CLEARED", "CHQ-BOUNCED", "CHQ-CANCELLED",
        })
        self.assertEqual(rows["CHQ-PENDING"][7:11], [
            "Bank Muscat", self.issue_date, self.payment_date, "Pending",
        ])
        self.assertEqual(rows["CHQ-CLEARED"][10], "Cleared")
        self.assertEqual(rows["CHQ-BOUNCED"][10], "Bounced")
        self.assertEqual(rows["CHQ-CANCELLED"][10], "Cancelled")
        self.assertEqual(len(rows), 4)

    def test_row_totals_and_summary_use_locked_calculations(self):
        data = self.report_data()
        purchase_rows = [row for row in data["rows"] if row[2] == self.purchase.purchase_number]
        self.assertEqual(len(purchase_rows), 7)
        for row in purchase_rows:
            self.assertEqual(row[11:15], [
                D("350.000"), D("100.000"), D("550.000"), D("650.000"),
            ])
        self.assertIsNone(data["totals"])
        self.assertEqual(dict(data["summary"]), {
            "Total Payment Amount": D("640.000"),
            "Total Pending Cheques": D("100.000"),
            "Total Cleared Cheques": D("125.000"),
        })
        self.assertNotIn("Supplier Balance", dict(data["summary"]))

    def test_supplier_method_and_cheque_status_filters(self):
        cases = [
            ({"supplier": self.other_supplier.pk}, 1, self.other_supplier.name),
            ({"method": SupplierPayment.Method.BANK}, 1, "Bank Transfer"),
            ({"cheque_status": SupplierPayment.ChequeStatus.BOUNCED}, 1, "Bounced"),
        ]
        for params, count, expected in cases:
            with self.subTest(params=params):
                data = self.report_data(**params)
                self.assertEqual(len(data["rows"]), count)
                self.assertIn(expected, data["rows"][0])

    def test_branch_and_warehouse_filters(self):
        branch_data = self.report_data(branch=self.other_branch.pk)
        warehouse_data = self.report_data(warehouse=self.other_warehouse.pk)
        self.assertEqual(len(branch_data["rows"]), 1)
        self.assertEqual(len(warehouse_data["rows"]), 1)
        self.assertEqual(branch_data["rows"][0][2], self.other_purchase.purchase_number)
        self.assertEqual(warehouse_data["rows"][0][2], self.other_purchase.purchase_number)

    def test_date_filter_uses_business_local_payment_date(self):
        previous_local_day = datetime(2026, 7, 15, 19, 30, tzinfo=UTC)
        selected_local_day = datetime(2026, 7, 15, 21, 30, tzinfo=UTC)
        SupplierPayment.objects.for_business(self.business_a).update(
            created_at=previous_local_day,
        )
        SupplierPayment.objects.filter(pk=self.payments["bank"].pk).update(
            created_at=selected_local_day,
        )
        data = self.report_data(**{"from": "2026-07-16", "to": "2026-07-16"})
        self.assertEqual(len(data["rows"]), 1)
        self.assertEqual(data["rows"][0][0].isoformat(), "2026-07-16")
        self.assertEqual(data["rows"][0][4], "Bank Transfer")

    def test_tenant_data_and_filter_options_are_isolated(self):
        supplier_b = Supplier.objects.create(
            business=self.business_b, code="RPT-B", name="Other Tenant Payment Supplier",
        )
        purchase_b = purchases.create_purchase(
            business=self.business_b,
            supplier=supplier_b,
            branch=self.branch_b,
            warehouse=self.warehouse_b,
            rows=[{
                "product": self.product_b,
                "variant": None,
                "quantity": D("10"),
                "unit_cost": D("5"),
            }],
            user=self.owner_b,
            purchase_date=business_localdate(self.business_b),
        )
        purchases.record_purchase_payments(
            purchase=purchase_b,
            rows=[{"method": SupplierPayment.Method.CASH, "amount": D("10")}],
            user=self.owner_b,
        )

        response = self.client.get(self.report_url())
        self.assertNotContains(response, supplier_b.name)
        self.assertNotIn(supplier_b, list(response.context["suppliers"]))
        for params in (
            {"supplier": supplier_b.pk},
            {"branch": self.branch_b.pk},
            {"warehouse": self.warehouse_b.pk},
        ):
            with self.subTest(params=params):
                self.assertEqual(self.report_data(**params)["rows"], [])

    def test_unlinked_legacy_supplier_payment_is_one_safe_row(self):
        legacy = SupplierPayment.objects.create(
            business=self.business_a,
            payment_number="SPY-RPT-LEGACY",
            supplier=self.supplier,
            amount=D("20.000"),
            payment_method=self.cash_a,
            paid_by=self.owner_a,
        )
        row = next(row for row in self.report_data()["rows"] if row[5] == legacy.amount)
        self.assertEqual(row[2:5], [None, None, "Cash"])
        self.assertEqual(row[6:14], [None] * 8)
        self.assertEqual(row[14], self.supplier.balance)

    def test_html_and_browser_print_output(self):
        response = self.client.get(self.report_url(method=SupplierPayment.Method.BANK))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Supplier Payments &amp; Cheques")
        for column in self.columns:
            self.assertContains(response, column)
        self.assertContains(response, "Bank Transfer")
        self.assertContains(response, 'onclick="window.print()"')
        self.assertContains(response, 'name="supplier"')
        self.assertContains(response, 'name="cheque_status"')

    def test_csv_export_matches_columns_filters_and_neutralizes_formulas(self):
        self.supplier.name = "=2+2"
        self.supplier.save(update_fields=["name", "updated_at"])
        cheque = self.payments["CHQ-PENDING"]
        cheque.cheque_number = "@unsafe"
        cheque.bank_name = "+SUM(1,1)"
        cheque.save(update_fields=["cheque_number", "bank_name", "updated_at"])

        response = self.client.get(self.report_url(export="csv"))
        self.assertEqual(response.status_code, 200)
        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertEqual(rows[0], self.columns)
        self.assertIn("'=2+2", {row[1] for row in rows[1:9]})
        self.assertIn("'@unsafe", {row[6] for row in rows[1:9]})
        self.assertIn("'+SUM(1,1)", {row[7] for row in rows[1:9]})

        filtered = self.client.get(
            self.report_url(export="csv", method=SupplierPayment.Method.BANK),
        )
        filtered_rows = list(csv.reader(StringIO(filtered.content.decode("utf-8"))))
        self.assertEqual(filtered_rows[0], self.columns)
        self.assertEqual(len([row for row in filtered_rows[1:] if row]), 4)
        self.assertEqual(filtered_rows[1][4], "Bank Transfer")

    def test_xlsx_export_matches_columns_filters_and_preserves_decimal_input(self):
        self.supplier.name = "=2+2"
        self.supplier.save(update_fields=["name", "updated_at"])
        self.assertIsInstance(_cell(D("12345678901.123")), Decimal)
        response = self.client.get(
            self.report_url(export="xlsx", method=SupplierPayment.Method.BANK),
        )
        self.assertEqual(response.status_code, 200)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
        worksheet = workbook.active
        self.assertEqual(
            [worksheet.cell(row=1, column=index).value for index in range(1, 16)],
            self.columns,
        )
        self.assertEqual(worksheet.cell(row=2, column=2).value, "'=2+2")
        self.assertEqual(worksheet.cell(row=2, column=5).value, "Bank Transfer")
        self.assertEqual(worksheet.cell(row=2, column=6).value, 75)
        workbook.close()

    def test_pdf_export_uses_same_filtered_data_and_wide_layout(self):
        with patch("apps.reports.exports.render_pdf", return_value=b"%PDF-filtered") as pdf:
            response = self.client.get(
                self.report_url(export="pdf", method=SupplierPayment.Method.BANK),
            )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"%PDF"))
        context = pdf.call_args.args[1]
        self.assertEqual(context["data"]["columns"], self.columns)
        self.assertEqual(len(context["data"]["rows"]), 1)
        self.assertEqual(context["data"]["rows"][0][4], "Bank Transfer")
        html = render_to_string("reports/report_pdf.html", context)
        self.assertIn("@page { size: A3 landscape; margin: 10mm; }", html)
        for column in self.columns:
            self.assertIn(column, html)

        actual = self.client.get(
            self.report_url(export="pdf", method=SupplierPayment.Method.BANK),
        )
        self.assertEqual(actual.status_code, 200)
        self.assertEqual(actual["Content-Type"], "application/pdf")
        self.assertTrue(actual.content.startswith(b"%PDF"))

    def test_outstanding_supplier_balances_report_contract_is_unchanged(self):
        title, query, permission = REPORTS["supplier_balances"]
        self.assertEqual(title, "Outstanding supplier balances")
        self.assertIs(query, supplier_balances)
        self.assertEqual(permission, "reports.financial")
        data = query(self.business_a, {})
        self.assertEqual(data["columns"], ["Supplier", "Code", "Mobile", "Payable"])
        balances = {row[0]: row[3] for row in data["rows"]}
        self.assertEqual(balances[self.supplier.name], D("650.000"))
        self.assertEqual(balances[self.other_supplier.name], D("460.000"))
        self.assertEqual(data["totals"][3], D("1110.000"))
