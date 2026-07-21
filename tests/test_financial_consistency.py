"""Focused contracts for return-aware financial consistency."""
import csv
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.urls import reverse

from apps.branches.models import Branch, Warehouse
from apps.catalog.models import Product
from apps.core.date_ranges import business_localdate, business_timezone
from apps.customers.models import Customer
from apps.expenses.models import Expense, ExpenseCategory
from apps.reports.queries import (
    profit_loss,
    returns_report,
    sales_summary,
)
from apps.sales import financials
from apps.sales import services as sales
from apps.sales.models import Sale, SalePayment, SaleReturn

from .base import TenantTestCase

D = Decimal
ZERO = D("0.000")


class FinancialConsistencyTests(TenantTestCase):
    def setUp(self):
        self.allow_no_shift()
        self.client.force_login(self.owner_a)
        self.today = business_localdate(self.business_a)
        self.product_one = Product.objects.create(
            business=self.business_a,
            name="Financial Item A",
            sku="FIN-A",
            product_type=Product.Type.NON_STOCK,
            track_inventory=False,
            purchase_price=D("4.000"),
            sale_price=D("10.000"),
            tax_rate=self.tax_a,
        )
        self.product_two = Product.objects.create(
            business=self.business_a,
            name="Financial Item B",
            sku="FIN-B",
            product_type=Product.Type.NON_STOCK,
            track_inventory=False,
            purchase_price=D("7.000"),
            sale_price=D("20.000"),
            tax_rate=self.tax_a,
        )
        self.customer = Customer.objects.create(
            business=self.business_a,
            home_branch=self.branch_a,
            code="FIN-CUSTOMER",
            full_name="Financial Consistency Customer",
            credit_limit=D("1000.000"),
        )

    def make_two_item_sale(
        self,
        *,
        discounts=None,
        payments=None,
        branch=None,
        warehouse=None,
        customer=None,
    ):
        branch = branch or self.branch_a
        warehouse = warehouse or self.warehouse_a
        customer = customer or self.customer
        discounts = discounts or (ZERO, ZERO)
        items = [
            {
                "product": self.product_one,
                "quantity": D("1.000"),
                "unit_price": D("10.000"),
                "discount_amount": discounts[0],
            },
            {
                "product": self.product_two,
                "quantity": D("1.000"),
                "unit_price": D("20.000"),
                "discount_amount": discounts[1],
            },
        ]
        if branch == self.branch_a and warehouse == self.warehouse_a:
            return self.make_sale(
                items=items,
                payments=payments,
                customer=customer,
            )
        return sales.complete_sale(
            business=self.business_a,
            branch=branch,
            warehouse=warehouse,
            cashier=self.owner_a,
            customer=customer,
            membership=self.membership_a(),
            items=items,
            payments=payments
            or [{"method": self.cash_a, "amount": D("31.500")}],
        )

    def return_items(self, sale, indexes, *, method=SaleReturn.RefundMethod.CASH):
        items = list(sale.items.order_by("pk"))
        return sales.process_return(
            sale=sale,
            items=[
                {
                    "sale_item": items[index],
                    "quantity": D("1.000"),
                    "restock": False,
                }
                for index in indexes
            ],
            refund_method=method,
            user=self.owner_a,
            restock=False,
        )

    def set_sale_activity_date(self, sale, value):
        timestamp = datetime.combine(
            value,
            time(hour=12),
            tzinfo=business_timezone(self.business_a),
        )
        Sale.objects.filter(pk=sale.pk).update(sale_date=timestamp)
        SalePayment.objects.filter(sale=sale).update(payment_date=value)

    def set_return_activity_date(self, sale_return, value):
        timestamp = datetime.combine(
            value,
            time(hour=12),
            tzinfo=business_timezone(self.business_a),
        )
        SaleReturn.objects.filter(pk=sale_return.pk).update(created_at=timestamp)

    def filters(self, *, day=None, branch_id=None, warehouse_id=None):
        return {
            "date_from": day,
            "date_to": day,
            "branch_id": branch_id,
            "warehouse_id": warehouse_id,
            "allowed_branch_ids": self.membership_a().allowed_branch_ids,
        }

    def pnl_rows(self, *, day=None, branch_id=None, warehouse_id=None):
        data = profit_loss(
            self.business_a,
            self.filters(
                day=day,
                branch_id=branch_id,
                warehouse_id=warehouse_id,
            ),
        )
        return {row[0]: row[1] for row in data["rows"] if row[0]}

    def test_partial_return_invoice_detail_uses_remaining_cost_and_profit(self):
        sale = self.make_two_item_sale()
        self.return_items(sale, [0])
        sale.refresh_from_db()

        response = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        )

        self.assertEqual(sale.total, D("31.500"))
        self.assertEqual(sale.returned_amount, D("10.500"))
        self.assertEqual(sale.net_total, D("21.000"))
        self.assertEqual(response.context["net_cost"], D("7.000"))
        self.assertEqual(response.context["net_gross_profit"], D("13.000"))
        self.assertEqual(
            [item.display_net_profit for item in response.context["items"]],
            [ZERO, D("13.000")],
        )
        self.assertContains(response, "Net cost")
        self.assertContains(response, "Net gross profit")

    def test_full_return_invoice_zeroes_net_total_cost_and_profit(self):
        sale = self.make_two_item_sale()
        self.return_items(sale, [0, 1])
        sale.refresh_from_db()

        response = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        )

        self.assertEqual(sale.net_total, ZERO)
        self.assertEqual(response.context["net_cost"], ZERO)
        self.assertEqual(response.context["net_gross_profit"], ZERO)

    def test_no_return_invoice_preserves_original_cost_and_profit(self):
        sale = self.make_two_item_sale()

        response = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        )

        self.assertEqual(response.context["net_cost"], sale.total_cost)
        self.assertEqual(
            response.context["net_gross_profit"], sale.gross_profit
        )
        self.assertEqual(sale.total_cost, D("11.000"))
        self.assertEqual(sale.gross_profit, D("19.000"))
        self.assertContains(response, ">Cost</th>", html=False)
        self.assertContains(response, ">Gross profit</th>", html=False)
        self.assertNotContains(response, "Net cost")

    def test_dashboard_year_cards_reconcile_all_financial_values(self):
        sale = self.make_two_item_sale(
            payments=[
                {"method": self.cash_a, "amount": D("21.000")},
                {"method": self.credit_a, "amount": D("10.500")},
            ]
        )
        self.return_items(sale, [0])
        category = ExpenseCategory.objects.create(
            business=self.business_a,
            name="Financial UAT Expense",
        )
        Expense.objects.create(
            business=self.business_a,
            expense_number="FIN-EXP-001",
            expense_date=self.today,
            branch=self.branch_a,
            category=category,
            amount=D("3.000"),
            status=Expense.Status.APPROVED,
            created_by=self.owner_a,
        )

        response = self.client.get(reverse("dashboard"))
        yearly = response.context["current_year"]

        self.assertEqual(yearly["gross_sales"], D("31.500"))
        self.assertEqual(yearly["total_returns"], D("10.500"))
        self.assertEqual(yearly["net_sales"], D("21.000"))
        self.assertEqual(yearly["total_income"], D("10.500"))
        self.assertEqual(yearly["net_income"], D("10.500"))
        self.assertEqual(yearly["total_receivable"], D("10.500"))
        self.assertEqual(yearly["gross_profit"], D("13.000"))
        self.assertEqual(yearly["estimated_net_profit"], D("10.000"))
        self.assertContains(
            response, "Gross Sales &mdash; This Year", html=True
        )
        self.assertContains(
            response, "Net Income Received &mdash; This Year", html=True
        )
        self.assertNotContains(response, "Total Sales &mdash; This Year")

    def test_profit_loss_partial_return_reverses_all_item_economics(self):
        sale = self.make_two_item_sale(
            discounts=(D("2.000"), D("4.000"))
        )
        self.return_items(sale, [0])

        rows = self.pnl_rows(day=self.today)
        sale.refresh_from_db()
        invoice_activity = financials.item_financial_summary_for_sale(sale)

        self.assertEqual(invoice_activity.net.tax, D("0.800"))
        self.assertEqual(
            invoice_activity.net.revenue_excluding_tax, D("16.000")
        )
        self.assertEqual(rows["Revenue (net of tax)"], D("20.000"))
        self.assertEqual(rows["Sales discounts given"], D("-4.000"))
        self.assertEqual(rows["Cost of goods sold"], D("-7.000"))
        self.assertEqual(rows["GROSS PROFIT"], D("9.000"))

        daily = sales_summary(self.business_a, self.filters(day=self.today))
        daily_row = next(row for row in daily["rows"] if row[1] == sale.invoice_number)
        self.assertEqual(daily_row[2], D("16.800"))
        self.assertEqual(daily_row[7], D("4.000"))
        self.assertEqual(daily_row[8], D("0.800"))
        self.assertEqual(daily_row[9], D("9.000"))
        returned = returns_report(self.business_a, self.filters(day=self.today))
        self.assertEqual(returned["totals"][9], D("8.400"))
        dashboard = self.client.get(
            reverse("dashboard"),
            {"from": self.today, "to": self.today},
        )
        self.assertEqual(dashboard.context["kpis"]["gross_profit"], D("9.000"))

    def test_profit_loss_return_is_scoped_to_originating_branch_and_warehouse(self):
        sale_a = self.make_two_item_sale()
        self.return_items(sale_a, [0])
        branch_b = Branch.objects.create(
            business=self.business_a,
            name="Financial Branch B",
            code="FIN-BR-B",
        )
        warehouse_b = Warehouse.objects.create(
            business=self.business_a,
            branch=branch_b,
            name="Financial Warehouse B",
            code="FIN-WH-B",
        )
        self.membership_a().branches.add(self.branch_a, branch_b)
        customer_b = Customer.objects.create(
            business=self.business_a,
            home_branch=branch_b,
            code="FIN-CUSTOMER-B",
            full_name="Financial Branch B Customer",
        )
        self.make_two_item_sale(
            branch=branch_b,
            warehouse=warehouse_b,
            customer=customer_b,
        )
        category = ExpenseCategory.objects.create(
            business=self.business_a,
            name="Branch-scoped Financial Expense",
        )
        Expense.objects.create(
            business=self.business_a,
            expense_number="FIN-BR-EXP-A",
            expense_date=self.today,
            branch=self.branch_a,
            category=category,
            amount=D("3.000"),
            status=Expense.Status.APPROVED,
            created_by=self.owner_a,
        )
        Expense.objects.create(
            business=self.business_a,
            expense_number="FIN-BR-EXP-B",
            expense_date=self.today,
            branch=branch_b,
            category=category,
            amount=D("5.000"),
            status=Expense.Status.APPROVED,
            created_by=self.owner_a,
        )

        all_rows = self.pnl_rows(day=self.today)
        branch_a_rows = self.pnl_rows(
            day=self.today,
            branch_id=self.branch_a.id,
        )
        branch_b_rows = self.pnl_rows(day=self.today, branch_id=branch_b.id)
        warehouse_b_rows = self.pnl_rows(
            day=self.today,
            warehouse_id=warehouse_b.id,
        )

        self.assertEqual(all_rows["GROSS PROFIT"], D("32.000"))
        self.assertEqual(branch_a_rows["GROSS PROFIT"], D("13.000"))
        self.assertEqual(branch_b_rows["GROSS PROFIT"], D("19.000"))
        self.assertEqual(warehouse_b_rows["GROSS PROFIT"], D("19.000"))
        self.assertEqual(all_rows["ESTIMATED NET PROFIT"], D("24.000"))
        self.assertEqual(
            branch_a_rows["ESTIMATED NET PROFIT"], D("10.000")
        )
        self.assertEqual(
            branch_b_rows["ESTIMATED NET PROFIT"], D("14.000")
        )
        self.assertEqual(
            warehouse_b_rows["ESTIMATED NET PROFIT"], D("14.000")
        )

    def test_persisted_refund_rounding_reconciles_every_profit_output(self):
        sale = self.make_sale(
            customer=self.customer,
            items=[{
                "product": self.product_one,
                "quantity": D("3.000"),
                "unit_price": D("0.016"),
            }],
        )
        sale_return = sales.process_return(
            sale=sale,
            items=[{
                "sale_item": sale.items.get(),
                "quantity": D("2.000"),
                "restock": False,
            }],
            refund_method=SaleReturn.RefundMethod.CASH,
            user=self.owner_a,
            restock=False,
        )
        sale.refresh_from_db()

        self.assertEqual(sale.total, D("0.050"))
        self.assertEqual(sale_return.refund_amount, D("0.034"))
        self.assertEqual(sale.net_total, D("0.016"))
        rows = self.pnl_rows(day=self.today)
        self.assertEqual(rows["Revenue (net of tax)"], D("0.015"))
        self.assertEqual(rows["Cost of goods sold"], D("-4.000"))
        self.assertEqual(rows["GROSS PROFIT"], D("-3.985"))

        detail = self.client.get(
            reverse("sales:detail", args=[sale.public_id])
        )
        self.assertEqual(detail.context["net_cost"], D("4.000"))
        self.assertEqual(detail.context["net_gross_profit"], D("-3.985"))
        returned = returns_report(self.business_a, self.filters(day=self.today))
        self.assertEqual(returned["totals"][9], D("0.034"))
        dashboard = self.client.get(
            reverse("dashboard"),
            {"from": self.today, "to": self.today},
        )
        self.assertEqual(dashboard.context["kpis"]["period_sales"], D("0.016"))
        self.assertEqual(
            dashboard.context["kpis"]["gross_profit"], D("-3.985")
        )
        self.assertEqual(
            dashboard.context["current_year"]["gross_profit"],
            D("-3.985"),
        )

    def test_profit_loss_books_return_on_actual_return_date(self):
        yesterday = self.today - timedelta(days=1)
        sale = self.make_two_item_sale()
        self.set_sale_activity_date(sale, yesterday)
        sale_return = self.return_items(sale, [0])
        self.set_return_activity_date(sale_return, self.today)

        sale_day = self.pnl_rows(day=yesterday)
        return_day = self.pnl_rows(day=self.today)

        self.assertEqual(sale_day["Revenue (net of tax)"], D("30.000"))
        self.assertEqual(sale_day["Cost of goods sold"], D("-11.000"))
        self.assertEqual(sale_day["GROSS PROFIT"], D("19.000"))
        self.assertEqual(return_day["Revenue (net of tax)"], D("-10.000"))
        self.assertEqual(return_day["Cost of goods sold"], D("4.000"))
        self.assertEqual(return_day["GROSS PROFIT"], D("-6.000"))

        sale_day_dashboard = self.client.get(
            reverse("dashboard"),
            {"from": yesterday, "to": yesterday},
        )
        return_day_dashboard = self.client.get(
            reverse("dashboard"),
            {"from": self.today, "to": self.today},
        )
        self.assertEqual(
            sale_day_dashboard.context["kpis"]["gross_profit"], D("19.000")
        )
        self.assertEqual(
            return_day_dashboard.context["kpis"]["gross_profit"], D("-6.000")
        )
        self.assertEqual(
            return_day_dashboard.context["kpis"]["period_sales"], D("-10.500")
        )

    def test_profit_loss_screen_print_and_exports_share_exact_dataset(self):
        sale = self.make_two_item_sale(
            discounts=(D("2.000"), D("4.000"))
        )
        self.return_items(sale, [0])
        url = reverse("reports:view", args=["profit_loss"])
        params = {
            "from": str(self.today),
            "to": str(self.today),
            "branch": str(self.branch_a.id),
        }

        screen = self.client.get(url, params)
        expected = screen.context["data"]
        self.assertContains(screen, "window.print()")

        csv_response = self.client.get(url, {**params, "export": "csv"})
        csv_rows = list(
            csv.reader(StringIO(csv_response.content.decode("utf-8-sig")))
        )
        expected_csv_rows = [expected["columns"]] + [
            ["-" if value in (None, "") else str(value) for value in row]
            for row in expected["rows"]
        ]
        self.assertEqual(csv_rows, expected_csv_rows)

        from openpyxl import load_workbook

        xlsx_response = self.client.get(url, {**params, "export": "xlsx"})
        workbook = load_workbook(
            BytesIO(xlsx_response.content),
            data_only=True,
            read_only=True,
        )
        xlsx_rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(list(xlsx_rows[0]), expected["columns"])
        for actual, row in zip(xlsx_rows[1:], expected["rows"], strict=True):
            self.assertEqual(actual[0], "-" if not row[0] else row[0])
            if row[1] in (None, ""):
                self.assertEqual(actual[1], "-")
            else:
                self.assertEqual(
                    D(str(actual[1])).quantize(D("0.001")),
                    row[1],
                )

        with patch(
            "apps.reports.exports.render_pdf",
            return_value=b"%PDF financial consistency",
        ) as render_pdf:
            pdf_response = self.client.get(url, {**params, "export": "pdf"})
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(render_pdf.call_args.args[1]["data"], expected)
