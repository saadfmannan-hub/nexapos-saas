import csv
import json
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import RequestFactory
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.catalog.models import Product
from apps.reports.queries import sales_detailed
from apps.sales import services as sales
from apps.sales.models import PaymentMethod, Sale, SaleItem
from apps.sales.views import _invoice_context, _job_card_context

from .base import TenantTestCase

D = Decimal


class PosCollectionTypeTests(TenantTestCase):
    def setUp(self):
        self.allow_no_shift()
        self.product_a.is_tailoring_item = True
        self.product_a.save(update_fields=["is_tailoring_item"])
        self.client.force_login(self.owner_a)

    def payload(self, collection_type="normal", **overrides):
        payload = {
            "branch_id": self.branch_a.id,
            "customer_id": self.walk_in_a.id,
            "items": [{
                "product_id": self.product_a.id,
                "variant_id": None,
                "quantity": "1",
                "unit_price": "10.000",
                "discount_amount": "0",
                "garment_classification": "adult",
                "collection_type": collection_type,
                "tailoring_details": {},
            }],
            "payments": [{"method_id": self.cash_a.id, "amount": "10.500"}],
            "invoice_discount": "0",
            "delivery_date": str(timezone.localdate()),
            "priority": "normal",
        }
        payload.update(overrides)
        return payload

    def checkout(self, payload):
        return self.client.post(
            reverse("sales:pos_checkout"),
            json.dumps(payload),
            content_type="application/json",
        )

    def sale_for(self, response):
        data = response.json()
        self.assertTrue(data["ok"], data)
        return Sale.objects.for_business(self.business_a).get(
            public_id=data["sale"]["public_id"]
        )

    def tailoring_sale(self, collection_type="premium"):
        return self.make_sale(
            items=[{
                "product": self.product_a,
                "quantity": D("1"),
                "unit_price": D("10"),
                "garment_classification": "adult",
                "collection_type": collection_type,
            }],
            delivery_date=timezone.localdate(),
        )

    def test_pos_defaults_tailoring_lines_to_normal_and_hides_retail_selector(self):
        html = self.client.get(reverse("sales:pos")).content.decode()
        self.assertIn("collection_type: p.is_tailoring_item ? 'normal' : ''", html)
        self.assertIn('x-show="line.is_tailoring_item"', html)
        self.assertIn('value="normal"', html)
        self.assertIn('value="premium"', html)

    def test_premium_is_stored_per_sale_item(self):
        sale = self.sale_for(self.checkout(self.payload("premium")))
        self.assertEqual(sale.items.get().collection_type, "premium")
        self.assertEqual(sale.items.get().collection_type_label, "Premium")

    def test_mixed_invoice_stores_normal_and_premium_per_line(self):
        payload = self.payload()
        premium = dict(payload["items"][0])
        premium["collection_type"] = "premium"
        premium["garment_classification"] = "child"
        payload["items"].append(premium)
        payload["payments"][0]["amount"] = "21.000"
        sale = self.sale_for(self.checkout(payload))
        self.assertEqual(
            list(sale.items.order_by("id").values_list("collection_type", flat=True)),
            ["normal", "premium"],
        )

    def test_checkout_rejects_missing_or_invalid_tailoring_collection(self):
        missing = self.payload()
        missing["items"][0].pop("collection_type")
        cases = (("missing", missing), ("invalid", self.payload("vip")))
        for name, payload in cases:
            with self.subTest(name=name):
                data = self.checkout(payload).json()
                self.assertFalse(data["ok"])
                self.assertEqual(
                    data["errors"]["items.0.collection_type"],
                    "Select Normal or Premium for every garment.",
                )

    def test_retail_does_not_require_collection(self):
        self.product_a.is_tailoring_item = False
        self.product_a.save(update_fields=["is_tailoring_item"])
        payload = self.payload("", delivery_date=None)
        payload["items"][0]["garment_classification"] = ""
        sale = self.sale_for(self.checkout(payload))
        self.assertEqual(sale.items.get().collection_type, "")

    def test_held_cart_round_trip_preserves_premium_and_restore_logic(self):
        cart = {"items": self.payload("premium")["items"], "priority": "normal"}
        response = self.client.post(
            reverse("sales:pos_hold"),
            json.dumps({
                "branch_id": self.branch_a.id,
                "label": "Premium garment",
                "cart": cart,
            }),
            content_type="application/json",
        )
        self.assertTrue(response.json()["ok"])
        held = self.client.get(reverse("sales:pos_held_list")).json()["held"][0]
        self.assertEqual(held["cart"]["items"][0]["collection_type"], "premium")
        html = self.client.get(reverse("sales:pos")).content.decode()
        self.assertIn("['normal', 'premium'].includes(collectionType)", html)
        self.assertIn("line.is_tailoring_item ? 'normal' : ''", html)

    def test_line_discount_ui_and_pos_workflow_are_removed(self):
        html = self.client.get(reverse("sales:pos")).content.decode()
        self.assertNotIn('title="Line discount"', html)
        self.assertNotIn('x-model="line.discount_amount"', html)
        self.assertIn("discount_amount: 0", html)

    def test_pos_ignores_forged_line_discount_but_invoice_discount_still_works(self):
        payload = self.payload(invoice_discount="1.000")
        payload["items"][0]["discount_amount"] = "9.000"
        payload["payments"][0]["amount"] = "9.450"
        sale = self.sale_for(self.checkout(payload))
        item = sale.items.get()
        self.assertEqual(item.discount_amount, D("0"))
        self.assertEqual(sale.discount_amount, D("1.000"))
        self.assertEqual(sale.tax_amount, D("0.450"))
        self.assertEqual(sale.total, D("9.450"))

    def test_historical_blank_collection_remains_valid_and_readable(self):
        sale = self.tailoring_sale()
        item = sale.items.get()
        item.collection_type = ""
        item.save(update_fields=["collection_type"])
        item.refresh_from_db()
        self.assertEqual(item.collection_type_label, "Legacy / Not Recorded")

    def test_legacy_service_call_without_collection_defaults_to_normal(self):
        sale = self.make_sale(
            items=[{
                "product": self.product_a,
                "quantity": D("1"),
                "unit_price": D("10"),
                "garment_classification": "adult",
            }],
            delivery_date=timezone.localdate(),
        )
        self.assertEqual(sale.items.get().collection_type, "normal")

    def test_sale_detail_and_all_invoice_formats_show_collection(self):
        sale = self.tailoring_sale()
        detail = self.client.get(reverse("sales:detail", args=[sale.public_id]))
        self.assertContains(detail, "Collection:")
        self.assertContains(detail, "Premium")
        a4 = self.client.get(reverse("sales:invoice", args=[sale.public_id]))
        self.assertContains(a4, "Collection: Premium")
        receipt_80 = self.client.get(reverse("sales:receipt", args=[sale.public_id]))
        self.assertContains(receipt_80, "Collection: Premium")
        receipt_58 = render_to_string("invoices/receipt_58mm.html", _invoice_context(sale))
        self.assertIn("Collection: Premium", receipt_58)

    def test_workshop_job_card_shows_collection(self):
        sale = self.tailoring_sale()
        item = sale.items.select_related("product__unit", "variant").get()
        request = RequestFactory().get("/")
        request.business = self.business_a
        html = render_to_string(
            "invoices/workshop_job_card.html",
            _job_card_context(sale, request, [item], sale_item=item),
        )
        self.assertIn('<div class="label">Collection</div>', html)
        self.assertIn('<div class="value">Premium</div>', html)

    def test_detailed_report_and_exports_show_collection(self):
        self.tailoring_sale()
        data = sales_detailed(self.business_a, {})
        collection_index = data["columns"].index("Collection")
        self.assertEqual(data["rows"][0][collection_index], "Premium")
        response = self.client.get(
            reverse("reports:view", args=["sales_detailed"]), {"export": "csv"}
        )
        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertIn("Collection", rows[0])
        self.assertEqual(rows[1][rows[0].index("Collection")], "Premium")

        xlsx_response = self.client.get(
            reverse("reports:view", args=["sales_detailed"]), {"export": "xlsx"}
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
        xlsx_rows = list(workbook.active.iter_rows(values_only=True))
        xlsx_collection_index = list(xlsx_rows[0]).index("Collection")
        self.assertEqual(xlsx_rows[1][xlsx_collection_index], "Premium")

        with patch("apps.reports.exports.render_pdf", return_value=b"%PDF fake") as render_pdf:
            pdf_response = self.client.get(
                reverse("reports:view", args=["sales_detailed"]), {"export": "pdf"}
            )
        self.assertEqual(pdf_response.status_code, 200)
        pdf_data = render_pdf.call_args.args[1]["data"]
        pdf_collection_index = pdf_data["columns"].index("Collection")
        self.assertEqual(pdf_data["rows"][0][pdf_collection_index], "Premium")

    def test_collection_remains_tenant_scoped(self):
        self.tailoring_sale()
        product_b = Product.objects.for_business(self.business_b).get(
            pk=self.product_b.pk
        )
        product_b.is_tailoring_item = True
        product_b.save(update_fields=["is_tailoring_item"])
        self.allow_no_shift(self.business_b)
        cash_b = PaymentMethod.objects.for_business(self.business_b).get(kind="cash")
        sales.complete_sale(
            business=self.business_b,
            branch=self.branch_b,
            warehouse=self.warehouse_b,
            cashier=self.owner_b,
            customer=self.walk_in_b,
            membership=self.business_b.memberships.get(user=self.owner_b),
            items=[{
                "product": product_b,
                "quantity": D("1"),
                "unit_price": product_b.sale_price,
                "garment_classification": "child",
                "collection_type": "normal",
            }],
            payments=[{"method": cash_b, "amount": D("5.250")}],
            delivery_date=timezone.localdate(),
        )
        data = sales_detailed(self.business_a, {})
        self.assertNotIn(product_b.name, [row[5] for row in data["rows"]])

    def test_adult_child_and_fabric_values_are_unchanged(self):
        self.product_a.estimated_adult_fabric = D("3.500")
        self.product_a.save(update_fields=["estimated_adult_fabric"])
        sale = self.sale_for(self.checkout(self.payload("premium")))
        item = sale.items.get()
        self.assertEqual(
            item.garment_classification,
            SaleItem.GarmentClassification.ADULT,
        )
        self.assertEqual(item.estimated_fabric, D("3.500"))
