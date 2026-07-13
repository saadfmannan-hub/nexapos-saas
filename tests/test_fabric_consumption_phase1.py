import csv
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.template.loader import render_to_string
from django.test import RequestFactory
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import Membership, Role, User
from apps.api.serializers import ProductSerializer, SaleSerializer
from apps.branches.models import Branch, Warehouse
from apps.catalog.forms import ProductForm
from apps.catalog.models import Product
from apps.inventory.models import StockMovement
from apps.reports.queries import sales_detailed
from apps.sales import services as sales
from apps.sales.models import PaymentMethod, Sale, SaleItem
from apps.sales.services import SaleError
from apps.sales.views import _job_card_context

from .base import TenantTestCase


D = Decimal


class FabricConsumptionPhase1Tests(TenantTestCase):
    def setUp(self):
        self.allow_no_shift()
        self.product_a.is_tailoring_item = True
        self.product_a.save(update_fields=["is_tailoring_item"])
        self.client.force_login(self.owner_a)

    def tailoring_sale(self, lines):
        return self.make_sale(
            items=[{
                "product": self.product_a,
                "quantity": D(str(quantity)),
                "unit_price": self.product_a.sale_price,
                "garment_classification": classification,
            } for classification, quantity in lines],
            delivery_date=timezone.localdate(),
        )

    def record_actual(self, item, value, *, user=None, membership=None):
        return sales.update_actual_fabric(
            sale_item=item,
            actual_fabric_used=value,
            user=user or self.owner_a,
            membership=membership or self.membership_a(),
        )

    def product_form_data(self, **overrides):
        data = {
            "name": "Configured Tailoring Garment",
            "product_type": Product.Type.NON_STOCK,
            "internal_code": "",
            "sku": "TAIL-CONFIG",
            "barcode": "",
            "purchase_price": "0.000",
            "sale_price": "10.000",
            "wholesale_price": "0.000",
            "minimum_sale_price": "0.000",
            "price_includes_tax": "",
            "reorder_level": "0.000",
            "allow_discount": "on",
            "is_tailoring_item": "on",
            "estimated_adult_fabric": "3.500",
            "estimated_child_fabric": "2.250",
            "description": "",
            "is_active": "on",
        }
        data.update(overrides)
        return data

    def workshop_member(self, *, branches=None):
        user = User.objects.create_user(
            email=f"workshop-{User.objects.count()}@example.com",
            password="StrongPass123!",
            full_name="Workshop User",
        )
        role = Role.objects.create(
            business=self.business_a,
            name=f"Authorized Workshop {Role.objects.count()}",
            permissions=["sales.view", "workshop.fabric_actual"],
        )
        membership = Membership.objects.create(
            business=self.business_a,
            user=user,
            role=role,
        )
        if branches:
            membership.branches.add(*branches)
        return user, membership

    def test_adult_estimate_is_quantity_times_configured_meters(self):
        item = self.tailoring_sale([("adult", 2)]).items.get()
        self.assertEqual(item.estimated_fabric, D("7.000"))
        self.assertIsNone(item.actual_fabric_used)
        self.assertIsNone(item.fabric_variance)

    def test_child_estimate_is_quantity_times_configured_meters(self):
        item = self.tailoring_sale([("child", 3)]).items.get()
        self.assertEqual(item.estimated_fabric, D("6.750"))

    def test_mixed_invoice_persists_each_line_estimate(self):
        items = self.tailoring_sale([("adult", 2), ("child", 3)]).items.order_by("id")
        self.assertEqual(
            list(items.values_list("estimated_fabric", flat=True)),
            [D("7.000"), D("6.750")],
        )

    def test_estimate_rounds_once_to_three_decimal_places(self):
        self.product_a.estimated_adult_fabric = D("2.345")
        self.product_a.save(update_fields=["estimated_adult_fabric"])
        item = self.make_sale(
            items=[{
                "product": self.product_a,
                "quantity": D("1.234"),
                "unit_price": self.product_a.sale_price,
                "garment_classification": "adult",
            }],
            payments=[{"method": self.cash_a, "amount": D("12.960")}],
            delivery_date=timezone.localdate(),
        ).items.get()
        self.assertEqual(item.estimated_fabric, D("2.894"))

    def test_product_edit_does_not_rewrite_historical_estimate(self):
        item = self.tailoring_sale([("adult", 1)]).items.get()
        self.product_a.estimated_adult_fabric = D("4.750")
        self.product_a.save(update_fields=["estimated_adult_fabric"])
        item.refresh_from_db()
        self.assertEqual(item.estimated_fabric, D("3.500"))

    def test_missing_selected_estimate_blocks_sale_without_partial_records(self):
        self.product_a.estimated_adult_fabric = None
        self.product_a.save(update_fields=["estimated_adult_fabric"])
        sales_before = Sale.objects.for_business(self.business_a).count()
        movements_before = StockMovement.objects.for_business(self.business_a).count()
        with self.assertRaisesMessage(SaleError, "Configure Estimated Adult Fabric"):
            self.tailoring_sale([("adult", 1)])
        self.assertEqual(Sale.objects.for_business(self.business_a).count(), sales_before)
        self.assertEqual(
            StockMovement.objects.for_business(self.business_a).count(),
            movements_before,
        )

    def test_historical_rows_remain_null_and_readable(self):
        self.product_a.is_tailoring_item = False
        self.product_a.save(update_fields=["is_tailoring_item"])
        item = self.make_sale().items.get()
        self.product_a.is_tailoring_item = True
        self.product_a.save(update_fields=["is_tailoring_item"])
        item.refresh_from_db()
        self.assertIsNone(item.estimated_fabric)
        self.assertIsNone(item.actual_fabric_used)
        self.assertIsNone(item.fabric_variance)

    def test_actual_fabric_and_variance_are_recorded_without_stock_movement(self):
        item = self.tailoring_sale([("adult", 1)]).items.get()
        movements_before = StockMovement.objects.for_business(self.business_a).count()
        updated = self.record_actual(item, "3.750")
        self.assertEqual(updated.actual_fabric_used, D("3.750"))
        self.assertEqual(updated.fabric_variance, D("0.250"))
        self.assertEqual(
            StockMovement.objects.for_business(self.business_a).count(),
            movements_before,
        )

    def test_actual_fabric_can_be_cleared_without_inventing_variance(self):
        item = self.tailoring_sale([("child", 1)]).items.get()
        item = self.record_actual(item, "2.000")
        item = self.record_actual(item, "")
        self.assertIsNone(item.actual_fabric_used)
        self.assertIsNone(item.fabric_variance)

    def test_non_stock_tailoring_workflow_creates_no_stock_movement(self):
        product = Product.objects.create(
            business=self.business_a,
            name="Made to Measure Garment",
            sku="MTM-NONSTOCK",
            product_type=Product.Type.NON_STOCK,
            track_inventory=False,
            sale_price=D("20.000"),
            is_tailoring_item=True,
            estimated_adult_fabric=D("3.500"),
            estimated_child_fabric=D("2.250"),
        )
        before = StockMovement.objects.for_business(self.business_a).count()
        sale = self.make_sale(
            items=[{
                "product": product,
                "quantity": D("1"),
                "unit_price": product.sale_price,
                "garment_classification": "adult",
            }],
            delivery_date=timezone.localdate(),
        )
        self.record_actual(sale.items.get(), "3.400")
        self.assertEqual(StockMovement.objects.for_business(self.business_a).count(), before)

    def test_product_form_validates_required_negative_invalid_and_max_values(self):
        valid = ProductForm(self.business_a, data=self.product_form_data())
        self.assertTrue(valid.is_valid(), valid.errors)
        for field, value in (
            ("estimated_adult_fabric", ""),
            ("estimated_adult_fabric", "-0.001"),
            ("estimated_child_fabric", "invalid"),
            ("estimated_child_fabric", "1000.001"),
        ):
            with self.subTest(field=field, value=value):
                form = ProductForm(
                    self.business_a,
                    data=self.product_form_data(**{field: value}),
                )
                self.assertFalse(form.is_valid())
                self.assertIn(field, form.errors)

    def test_non_tailoring_product_form_clears_fabric_defaults(self):
        data = self.product_form_data()
        data.pop("is_tailoring_item")
        form = ProductForm(self.business_a, data=data)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertIsNone(form.cleaned_data["estimated_adult_fabric"])
        self.assertIsNone(form.cleaned_data["estimated_child_fabric"])

    def test_product_form_shows_meter_controls_only_in_tailoring_section(self):
        response = self.client.get(
            reverse("catalog:product_edit", args=[self.product_a.public_id])
        )
        self.assertContains(response, 'x-show="isTailoring"')
        self.assertContains(response, "Estimated Adult Fabric (Meters)")
        self.assertContains(response, "Estimated Child Fabric (Meters)")

    def test_migrated_tailoring_product_shows_unconfigured_defaults_safely(self):
        self.product_a.estimated_adult_fabric = None
        self.product_a.estimated_child_fabric = None
        self.product_a.save(update_fields=[
            "estimated_adult_fabric",
            "estimated_child_fabric",
        ])
        response = self.client.get(
            reverse("catalog:product_detail", args=[self.product_a.public_id])
        )
        self.assertContains(response, "Not configured", count=2)

    def test_owner_and_authorized_workshop_user_can_update_actual(self):
        sale = self.tailoring_sale([("adult", 1)])
        item = sale.items.get()
        owner_response = self.client.post(
            reverse("sales:sale_item_update_fabric", args=[sale.public_id, item.id]),
            {"actual_fabric_used": "3.600"},
        )
        self.assertEqual(owner_response.status_code, 302)
        user, _membership = self.workshop_member()
        self.client.force_login(user)
        response = self.client.post(
            reverse("sales:sale_item_update_fabric", args=[sale.public_id, item.id]),
            {"actual_fabric_used": "3.550"},
        )
        self.assertEqual(response.status_code, 302)
        item.refresh_from_db()
        self.assertEqual(item.actual_fabric_used, D("3.550"))

    def test_cashier_cannot_update_actual_fabric(self):
        sale = self.tailoring_sale([("adult", 1)])
        item = sale.items.get()
        self.client.force_login(self.cashier_a)
        response = self.client.post(
            reverse("sales:sale_item_update_fabric", args=[sale.public_id, item.id]),
            {"actual_fabric_used": "3.500"},
        )
        self.assertEqual(response.status_code, 403)
        item.refresh_from_db()
        self.assertIsNone(item.actual_fabric_used)

    def test_negative_and_invalid_actual_values_are_rejected(self):
        sale = self.tailoring_sale([("adult", 1)])
        item = sale.items.get()
        for value in ("-1", "invalid"):
            with self.subTest(value=value):
                response = self.client.post(
                    reverse(
                        "sales:sale_item_update_fabric",
                        args=[sale.public_id, item.id],
                    ),
                    {"actual_fabric_used": value},
                )
                self.assertEqual(response.status_code, 302)
                item.refresh_from_db()
                self.assertIsNone(item.actual_fabric_used)

    def test_actual_update_is_tenant_isolated(self):
        self.allow_no_shift(self.business_b)
        self.product_b.is_tailoring_item = True
        self.product_b.save(update_fields=["is_tailoring_item"])
        cash_b = PaymentMethod.objects.for_business(self.business_b).get(kind="cash")
        sale_b = sales.complete_sale(
            business=self.business_b,
            branch=self.branch_b,
            warehouse=self.warehouse_b,
            cashier=self.owner_b,
            customer=self.walk_in_b,
            membership=self.business_b.memberships.get(user=self.owner_b),
            items=[{
                "product": self.product_b,
                "quantity": D("1"),
                "unit_price": self.product_b.sale_price,
                "garment_classification": "adult",
            }],
            payments=[{"method": cash_b, "amount": D("5.000")}],
            delivery_date=timezone.localdate(),
        )
        item_b = sale_b.items.get()
        response = self.client.post(
            reverse(
                "sales:sale_item_update_fabric",
                args=[sale_b.public_id, item_b.id],
            ),
            {"actual_fabric_used": "3.500"},
        )
        self.assertEqual(response.status_code, 404)

    def test_actual_update_is_branch_isolated(self):
        second_branch = Branch.objects.create(
            business=self.business_a,
            name="Workshop Restricted",
            code="WORKSHOP-RESTRICTED",
        )
        second_warehouse = Warehouse.objects.create(
            business=self.business_a,
            branch=second_branch,
            name="Workshop Restricted Warehouse",
            code="WORKSHOP-RESTRICTED-WH",
        )
        product = Product.objects.create(
            business=self.business_a,
            name="Restricted Garment",
            product_type=Product.Type.NON_STOCK,
            track_inventory=False,
            sale_price=D("10"),
            is_tailoring_item=True,
            estimated_adult_fabric=D("3.500"),
            estimated_child_fabric=D("2.250"),
        )
        sale = sales.complete_sale(
            business=self.business_a,
            branch=second_branch,
            warehouse=second_warehouse,
            cashier=self.owner_a,
            customer=self.walk_in_a,
            membership=self.membership_a(),
            items=[{
                "product": product,
                "quantity": D("1"),
                "unit_price": product.sale_price,
                "garment_classification": "adult",
            }],
            payments=[{"method": self.cash_a, "amount": D("10.000")}],
            delivery_date=timezone.localdate(),
        )
        user, _membership = self.workshop_member(branches=[self.branch_a])
        self.client.force_login(user)
        response = self.client.post(
            reverse(
                "sales:sale_item_update_fabric",
                args=[sale.public_id, sale.items.get().id],
            ),
            {"actual_fabric_used": "3.500"},
        )
        self.assertEqual(response.status_code, 403)

    def test_sale_detail_and_job_card_show_fabric_values(self):
        sale = self.tailoring_sale([("adult", 1)])
        item = self.record_actual(sale.items.get(), "3.750")
        detail = self.client.get(reverse("sales:detail", args=[sale.public_id]))
        self.assertContains(detail, "Estimated:")
        self.assertContains(detail, "3.500 m")
        self.assertContains(detail, "3.750 m")
        self.assertContains(detail, "0.250 m")
        self.assertContains(detail, 'name="actual_fabric_used"')

        request = RequestFactory().get("/")
        request.business = self.business_a
        html = render_to_string(
            "invoices/workshop_job_card.html",
            _job_card_context(sale, request, [item], sale_item=item),
        )
        self.assertIn("Estimated Fabric", html)
        self.assertIn("Actual Fabric", html)
        self.assertIn("Variance", html)
        self.assertIn("3.750 m", html)

    def test_cashier_sale_detail_does_not_render_actual_edit_control(self):
        sale = self.tailoring_sale([("adult", 1)])
        self.client.force_login(self.cashier_a)
        response = self.client.get(reverse("sales:detail", args=[sale.public_id]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="actual_fabric_used"')

    def test_report_has_fabric_columns_and_totals(self):
        sale = self.tailoring_sale([("adult", 1), ("child", 1)])
        items = list(sale.items.order_by("id"))
        self.record_actual(items[0], "3.750")
        self.record_actual(items[1], "2.000")
        data = sales_detailed(self.business_a, {})
        self.assertIn("Estimated Fabric", data["columns"])
        self.assertIn("Actual Fabric", data["columns"])
        self.assertIn("Variance", data["columns"])
        summary = dict(data["summary"])
        self.assertEqual(summary["Estimated Total"], D("5.750"))
        self.assertEqual(summary["Actual Total"], D("5.750"))
        self.assertEqual(summary["Variance Total"], D("0.000"))

    def test_csv_xlsx_and_pdf_exports_include_fabric_data(self):
        sale = self.tailoring_sale([("adult", 1)])
        self.record_actual(sale.items.get(), "3.750")
        url = reverse("reports:view", args=["sales_detailed"])

        csv_response = self.client.get(url, {"export": "csv"})
        csv_rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8"))))
        self.assertIn("Estimated Fabric", csv_rows[0])
        self.assertIn("Actual Fabric", csv_rows[0])
        self.assertIn("Variance", csv_rows[0])
        self.assertIn(["Estimated Total", "3.500"], csv_rows)
        self.assertIn(["Actual Total", "3.750"], csv_rows)
        self.assertIn(["Variance Total", "0.250"], csv_rows)

        xlsx_response = self.client.get(url, {"export": "xlsx"})
        workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
        xlsx_rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn("Estimated Fabric", xlsx_rows[0])
        self.assertIn("Actual Fabric", xlsx_rows[0])
        self.assertIn("Variance", xlsx_rows[0])

        with patch("apps.reports.exports.render_pdf", return_value=b"%PDF fake") as render_pdf:
            pdf_response = self.client.get(url, {"export": "pdf"})
        self.assertEqual(pdf_response.status_code, 200)
        pdf_data = render_pdf.call_args.args[1]["data"]
        self.assertIn("Estimated Fabric", pdf_data["columns"])
        self.assertEqual(dict(pdf_data["summary"])["Variance Total"], D("0.250"))

    def test_api_exposes_read_only_fabric_snapshots(self):
        sale = self.tailoring_sale([("child", 1)])
        item = self.record_actual(sale.items.get(), "2.000")
        product_data = ProductSerializer(self.product_a).data
        self.assertEqual(product_data["estimated_adult_fabric"], "3.500")
        self.assertEqual(product_data["estimated_child_fabric"], "2.250")
        item_data = SaleSerializer(sale).data["items"][0]
        self.assertEqual(item_data["estimated_fabric"], "2.250")
        self.assertEqual(item_data["actual_fabric_used"], "2.000")
        self.assertEqual(item_data["fabric_variance"], "-0.250")
        self.assertEqual(item.fabric_variance, D("-0.250"))

    def test_default_workshop_role_has_permission_and_cashier_does_not(self):
        workshop = Role.objects.for_business(self.business_a).get(name="Workshop Manager")
        cashier = Role.objects.for_business(self.business_a).get(name="Cashier")
        self.assertIn("workshop.fabric_actual", workshop.permissions)
        self.assertNotIn("workshop.fabric_actual", cashier.permissions)
